from decimal import Decimal
from pathlib import Path
import io
import zipfile

import pytest
from fastapi import HTTPException
from openpyxl import Workbook, load_workbook

import app.supplier_price_bridge as bridge_rules
from app.routers.supplier_price_bridge import _parse, _read_upload
from app.supplier_price_bridge import (
    canonical_price_digest,
    classify_deviation,
    collect_reason_codes,
    deviation_percent,
    normalize_price_cell,
    normalize_price,
    split_supplier_code,
    vat_details,
)
from app.auth import ROLE_PERMISSIONS


@pytest.mark.parametrize(
    ("raw", "amount", "currency"),
    [
        ("1 ,10 USD", Decimal("1.1000"), "USD"),
        ("€ 0 ,35", Decimal("0.3500"), "EUR"),
        ("600,-TL", Decimal("600.0000"), "TRY"),
        ("3.200,-TL", Decimal("3200.0000"), "TRY"),
        ("1.234,56", Decimal("1234.5600"), None),
    ],
)
def test_tampar_price_artifacts_are_normalized(raw, amount, currency):
    assert normalize_price(raw) == (amount, currency)


def test_digit_internal_spaces_are_compacted_without_breaking_separator_spacing():
    assert normalize_price("€ 2 0,00") == (Decimal("20.0000"), "EUR")
    assert normalize_price("€ 1 0,56") == (Decimal("10.5600"), "EUR")
    assert normalize_price("$ 1 ,10") == (Decimal("1.1000"), "USD")


@pytest.mark.parametrize("raw", ["2 0", "2 0%", "2 0 3,00", "20,00 30,00"])
def test_digit_internal_spaces_do_not_merge_non_price_or_multiple_values(raw):
    with pytest.raises(ValueError, match="Fiyat ayrıştırılamadı"):
        normalize_price(raw)


def test_tampar_multiline_code_preserves_description():
    assert split_supplier_code("64S20-3602\n64TC/20diş", "Bıçak") == (
        "64S20-3602",
        "64TC/20diş\nBıçak",
    )


def test_tampar_split_code_continuation_is_joined():
    assert split_supplier_code("17SCH-\n16500.04", "Zıpka") == (
        "17SCH-16500.04",
        "Zıpka",
    )


@pytest.mark.parametrize(
    ("raw", "note"),
    [
        ("sorunuz", "sorunuz"),
        ("stok mevcut-sorunuz", "stok mevcut-sorunuz"),
        ("kalmadı", "kalmadı"),
        ("STOK AZ", "stok az"),
    ],
)
def test_status_text_is_not_guessed_as_price(raw, note):
    cell = normalize_price_cell(raw)
    assert cell.amount is None
    assert cell.status_note == note
    assert cell.reason_code == "NO_PRICE_STATUS"


def test_status_rule_mutant_is_caught(monkeypatch):
    def assert_rule():
        assert normalize_price_cell("sorunuz").reason_code == "NO_PRICE_STATUS"

    assert_rule()
    monkeypatch.setattr(bridge_rules, "NO_PRICE_NOTES", ())
    with pytest.raises(AssertionError):
        assert_rule()


def test_multiple_eur_prices_require_review():
    cell = normalize_price_cell("PEŞİN 3,08EUR / 30 GÜN 3,19EUR")
    assert cell.amount is None
    assert cell.reason_code == "MULTIPLE_PRICES"


@pytest.mark.parametrize(
    ("has_price", "has_note", "primary"),
    [
        (True, False, None),
        (True, True, None),
        (False, False, "INVALID_PRICE"),
        (False, True, "NO_PRICE_STATUS"),
    ],
)
def test_price_status_note_combination_matrix(has_price, has_note, primary):
    price = normalize_price_cell("10,00 TL" if has_price else None)
    note = normalize_price_cell("STOK AZ") if has_note else None
    codes, reason = collect_reason_codes(price, price, note)
    assert reason == primary
    assert codes["status_note"] == ("NO_PRICE_STATUS" if has_note else None)


def test_reason_priority_preserves_every_cell_reason():
    cash = normalize_price_cell("bozuk")
    term = normalize_price_cell("sorunuz")
    note = normalize_price_cell("3,08EUR / 3,19EUR")
    codes, primary = collect_reason_codes(cash, term, note)
    assert codes == {
        "price_cash": "INVALID_PRICE",
        "price_term": "NO_PRICE_STATUS",
        "status_note": "MULTIPLE_PRICES",
    }
    assert primary == "MULTIPLE_PRICES"


def test_spaced_vat_included_is_normalized():
    assert vat_details("2 0 %\nKDV\nDAHİL") == (Decimal("20.00"), True)


def test_supplier_price_role_matrix_is_narrow():
    for role in ("admin", "yonetici", "muhasebe"):
        assert "supplier_prices.import" in ROLE_PERMISSIONS[role] or "*" in ROLE_PERMISSIONS[role]
        assert "supplier_prices.apply" in ROLE_PERMISSIONS[role] or "*" in ROLE_PERMISSIONS[role]
    assert "supplier_prices.view" in ROLE_PERMISSIONS["depo"]
    assert "supplier_prices.import" not in ROLE_PERMISSIONS["depo"]
    assert "supplier_prices.apply" not in ROLE_PERMISSIONS["depo"]


def test_real_tampar_slice_contains_and_normalizes_locked_examples():
    fixture = Path(__file__).parent / "fixtures" / "tampar_slice.xlsx"
    workbook = load_workbook(fixture, read_only=True, data_only=True, keep_links=False)
    nonempty_rows = sum(
        1
        for sheet in workbook.worksheets
        for row in sheet.iter_rows(values_only=True)
        if any(value not in (None, "") for value in row)
    )
    assert fixture.stat().st_size <= 200 * 1024
    assert nonempty_rows <= 200
    assert workbook["06_Bicaklar"]["A6"].value == "64S20-3602\n64TC/20diş"
    assert workbook["06_Bicaklar"]["F6"].value == "1 ,10 USD"
    assert workbook["06_Bicaklar"]["F10"].value == "€ 0 ,35"
    assert workbook["07_Zipka_Mastar"]["E20"].value == "600,-TL"
    assert normalize_price(workbook["06_Bicaklar"]["F6"].value) == (Decimal("1.1000"), "USD")
    assert normalize_price(workbook["06_Bicaklar"]["F10"].value) == (Decimal("0.3500"), "EUR")
    assert normalize_price(workbook["07_Zipka_Mastar"]["E20"].value) == (Decimal("600.0000"), "TRY")


def test_zip_bomb_stops_at_uncompressed_entry_limit(monkeypatch):
    # Modül-düzeyi `from ... import _validate_archive` süpürmede BAYAT kalır:
    # `test_security_audit_visibility.py:43-44,60-61` `app.*`yi silip yeniden
    # yükler; `monkeypatch.setattr("app.routers....MAX_ENTRY_BYTES")` YENİ
    # modüle yazar, bayat fonksiyon kendi `__globals__`ında 100MB görür ve
    # HTTPException ATMAZ (DID NOT RAISE). Aynı import nesnesine yaz.
    import app.routers.supplier_price_bridge as bridge_mod

    archive_bytes = io.BytesIO()
    with zipfile.ZipFile(archive_bytes, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("xl/worksheets/sheet1.xml", b"0" * 4096)
    monkeypatch.setattr(bridge_mod, "MAX_ENTRY_BYTES", 1024)
    with pytest.raises(HTTPException, match="açılmış veri sınırını"):
        bridge_mod._validate_archive(archive_bytes.getvalue())


@pytest.mark.asyncio
async def test_26_mb_upload_is_rejected_before_first_read():
    class Upload:
        def __init__(self):
            self.calls = 0
            self.size = 26 * 1024 * 1024

        async def read(self, _size):
            self.calls += 1
            return b"x"

    upload = Upload()
    with pytest.raises(HTTPException, match="25 MB"):
        await _read_upload(upload)
    assert upload.calls == 0


def _workbook_content(workbook):
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _fixed_profile():
    return {
        "supplier_id": 1,
        "sheet_selector": "Tum_Tablolar",
        "header_row_strategy": "fixed:1",
        "column_map": {"code": "Kod", "price_cash": "Peşin", "price_term": "Vadeli"},
        "currency_mode": "fixed:TRY",
        "vat_source": "fixed:20",
        "warning_threshold": Decimal("25"),
        "blocked_threshold": Decimal("100"),
    }


def test_51000_row_workbook_is_rejected_before_database_writes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Tum_Tablolar"
    sheet.append(["Kod", "Peşin", "Vadeli"])
    for row_number in range(51_000):
        sheet.append([f"KOD-{row_number}", "1,00 TL", "1,10 TL"])

    class NoDatabaseWrites:
        def execute(self, *_args, **_kwargs):
            raise AssertionError("Satır sınırı aşılırken DB çağrısı yapılmamalı")

    with pytest.raises(HTTPException, match="50.000 satır"):
        _parse(NoDatabaseWrites(), 1, 1, _workbook_content(workbook), _fixed_profile())


def test_201_sheet_workbook_is_rejected_before_database_writes():
    workbook = Workbook()
    workbook.active.title = "Tum_Tablolar"
    workbook.active.append(["Kod", "Peşin", "Vadeli"])
    for number in range(200):
        workbook.create_sheet(f"Ek_{number:03d}")

    class NoDatabaseWrites:
        def execute(self, *_args, **_kwargs):
            raise AssertionError("Sheet sınırı aşılırken DB çağrısı yapılmamalı")

    with pytest.raises(HTTPException, match="200 sheet"):
        _parse(NoDatabaseWrites(), 1, 1, _workbook_content(workbook), _fixed_profile())


@pytest.mark.parametrize(
    ("new", "expected"),
    [
        ("124.9", "OK"),
        ("125.0", "WARNING"),
        ("199.9", "WARNING"),
        ("200.0", "BLOCKED"),
    ],
)
def test_deviation_boundaries_are_exact(new, expected):
    deviation = deviation_percent(Decimal("100"), Decimal(new))
    assert classify_deviation(deviation, Decimal("25"), Decimal("100")) == expected


def test_canonical_digest_is_ordered_and_uses_literal_null():
    rows = [
        {"part_id": 2, "price_cash": None, "price_term": "2", "currency": "EUR", "vat_rate": "20"},
        {"part_id": 1, "price_cash": "1", "price_term": "1.5", "currency": "USD", "vat_rate": "10"},
    ]
    assert canonical_price_digest(rows) == canonical_price_digest(reversed(rows))
    assert canonical_price_digest(rows) == "9065a4ecdc6e944a79a3b417a47b2a213f40aa44b3ebfe22cbb69341f044edf6"


def test_financial_bridge_module_does_not_use_float_constructor():
    app_dir = Path(__file__).parents[1] / "app"
    sources = [
        app_dir / "supplier_price_bridge.py",
        app_dir / "routers" / "supplier_price_bridge.py",
    ]
    assert all("float(" not in source.read_text("utf-8") for source in sources)
