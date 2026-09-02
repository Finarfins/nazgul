"""Uygulama Kayıt Çizelgesi — iki okuma ucu.

Testlerin ağırlığı ALTI iddiada:

1. **FORMÜL ENJEKSİYONU.** `input_name`, `notes`, gerekçeler ve
   `safety_warning` SERBEST METİNDİR ve müfettişin AÇACAĞI hücreye gidiyor.
   `=`, `+`, `-`, `@` ile başlayan bir değer Excel'de FORMÜL olur. Test hem
   üreticinin ham satırını hem KAYDEDİLMİŞ dosyanın hücresini okuyor: yalnız
   ham satıra bakmak, `_safe_excel_cell`in çağrılmadığı bir regresyonu
   göremezdi.
2. **EN AZ BİR SÜZGEÇ.** Süzgeçsiz çağrı 400. Tarih aralığı TEK BAŞINA
   yetmez — "2020-2026 arası" bütün tarihçedir.
3. **KARTEZYEN ÇARPIM YOK.** İki girdili bir faaliyet + iki hasat olan bir
   sezonda hasat satırı 2 kalmalı, çarpmamalı. `parcel_timeline` ile aynı
   tuzak, aynı şekille kuruluyor.
4. **GİRDİSİZ FAALİYET SATIRI KAYBETMİYOR.** Toprak işleme ve sulamanın girdisi
   yoktur; `INNER JOIN` mantığıyla yazılmış bir üretici bu kayıtları
   ÇİZELGEDEN SİLERDİ.
5. **PARA SÜTUNU YOK.** Çizelge `unit_cost`/`total_cost`/`revenue_amount` ve
   donmuş oranları TAŞIMAZ. `farm.py`deki okuma yolu maskesinin ETRAFINDAN
   dolaşan ikinci bir yol olmadığını çiviliyor.
6. **ÇAPRAZ KİRACI.** Başka firmanın kimliğiyle çağrılan süzgeç BOŞ döner.

Ayrıca ÖLÇÜLEN bir nokta: gün İSTANBUL takvimine göre yazılır —
`farm.py::_yerel_gun` ile aynı kural. UTC gününe düşülseydi uygulamanın iki
parçası (bekleme süresi hesabı ve çizelge) hangi gün olduğu konusunda
ANLAŞMAZDI; test gece yarısı sınırındaki bir kaydı bilerek kuruyor.
"""
from __future__ import annotations

import os
import subprocess
import sys
from pathlib import Path

BACKEND = Path(__file__).resolve().parents[1]


def run_logbook_smoke(database_url: str) -> None:
    env = os.environ.copy()
    env["DATABASE_URL"] = database_url
    env["PYTHONPATH"] = str(BACKEND)
    completed = subprocess.run(
        [sys.executable, "-c", _SMOKE],
        cwd=BACKEND, env=env, capture_output=True, text=True, timeout=300,
    )
    assert completed.returncode == 0, completed.stdout + "\n" + completed.stderr


def test_producer_logbook_sqlite(tmp_path: Path) -> None:
    run_logbook_smoke(f"sqlite:///{(tmp_path / 'kayit-defteri.db').as_posix()}")


def test_filter_is_required_without_a_database() -> None:
    """Süzgeç kapısı SAF: veritabanına hiç gitmeden reddediyor."""
    import pytest

    from app.uretici_kayit_defteri import (
        EN_AZ_BIR_SUZGEC, DefterHatasi, suzgec_dogrula,
    )

    with pytest.raises(DefterHatasi):
        suzgec_dogrula({})
    # Tarih aralığı TEK BAŞINA süzgeç değildir.
    with pytest.raises(DefterHatasi):
        suzgec_dogrula({"date_from": "2026-01-01", "date_to": "2026-12-31"})
    for ad in EN_AZ_BIR_SUZGEC:
        assert suzgec_dogrula({ad: 1})[ad] == 1


def test_no_money_column_is_offered() -> None:
    """Para sütunu ÇİZELGEDE YOK — maskeyi dolanan ikinci yol olmasın."""
    from app.uretici_kayit_defteri import FAALIYET_BASLIKLARI, HASAT_BASLIKLARI

    basliklar = " ".join(FAALIYET_BASLIKLARI + HASAT_BASLIKLARI).lower()
    for yasak in ("maliyet", "tutar", "fiyat", "gelir", "ücret"):
        assert yasak not in basliklar, (yasak, basliklar)


def test_the_sheet_states_what_it_is_and_claims_nothing_official() -> None:
    """Sayfadaki not, resmî form İDDİASINI açıkça reddetmeli."""
    from app.uretici_kayit_defteri import CIZELGE_ADI, CIZELGE_NOTU

    assert "Resmî bir form değildir" in CIZELGE_NOTU
    assert "onaylanmamıştır" in CIZELGE_NOTU
    assert "yerine geçmez" in CIZELGE_NOTU
    # Ad da iddia taşımıyor: "defter" değil "çizelge".
    assert CIZELGE_ADI == "Uygulama Kayıt Çizelgesi"


_SMOKE = r'''
from decimal import Decimal
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app
from app.uretici_kayit_defteri import FAALIYET_BASLIKLARI, HASAT_BASLIKLARI

ADMIN_PW = 'Defter!12345'

# Excel'de FORMÜL olarak yorumlanacak baş karakter.
ZEHIR = '=1+1'

# Para sızıntısı sondası. DEĞER ÖZELLİKLE ayırt edici: ilk denemede '100.00'
# kullanılmıştı ve PostgreSQL ikizi bunu yakaladı — parselin 100.0000 dekarlık
# ALANI o dizeyi İÇERİYOR, yani sonda kendi kendine yanlış alarm veriyordu.
# (SQLite alanı '100' olarak verdiği için orada sessizce geçiyordu.)
BIRIM_MALIYET = '137.77'


def admin_headers(client):
    login = client.post('/api/auth/login', json={'username':'admin','password':'admin123'})
    assert login.status_code == 200, login.text
    body = login.json()
    h = {'Authorization':'Bearer '+body['access_token'],
         'X-Company-ID':str(body['companies'][0]['id'])}
    ch = client.post('/api/auth/change-password', headers=h,
                     json={'current_password':'admin123','new_password':ADMIN_PW})
    assert ch.status_code == 200, ch.text
    h['Authorization'] = 'Bearer '+ch.json()['access_token']
    return h


with TestClient(app) as client:
    h = admin_headers(client)

    ciftlik = client.post('/api/farms', headers=h,
                          json={'code':'d1','name':'Defter Çiftlik',
                                'city':'Konya','district':'Çumra'}).json()
    p1 = client.post('/api/farm-parcels', headers=h,
                     json={'farm_id':ciftlik['id'],'code':'dp1','name':'Birinci Parsel',
                           'area_decare':'100.0000','parcel_no':'12','block_no':'34',
                           'neighborhood':'Merkez'}).json()
    p2 = client.post('/api/farm-parcels', headers=h,
                     json={'farm_id':ciftlik['id'],'code':'dp2','name':'İkinci Parsel',
                           'area_decare':'50.0000'}).json()

    s1 = client.post('/api/crop-seasons', headers=h,
                     json={'parcel_id':p1['id'],'season_year':2026,'crop':'Buğday',
                           'variety':'Bezostaja','started_on':'2026-03-01',
                           'planted_area_decare':'80.0000'}).json()
    s2 = client.post('/api/crop-seasons', headers=h,
                     json={'parcel_id':p2['id'],'season_year':2026,'crop':'Mısır',
                           'started_on':'2026-03-01'}).json()

    # --- KARTEZYEN ÇARPIM TUZAĞINI KURAN ŞEKİL -----------------------------
    # TEK ilaçlama faaliyeti, İKİ girdi; ayrıca İKİ hasat.
    ilac = client.post('/api/field-activities', headers=h,
                       json={'season_id':s1['id'],'activity_type':'SPRAYING',
                             'performed_at':'2026-05-10T09:00:00+03:00',
                             'applied_area_decare':'80.0000',
                             'preharvest_interval_days':21,
                             'reentry_interval_days':2,
                             'notes':ZEHIR})
    assert ilac.status_code == 201, ilac.text
    for ad in (ZEHIR, 'Normal İlaç'):
        g = client.post(f"/api/field-activities/{ilac.json()['id']}/inputs", headers=h,
                        json={'input_name':ad,'quantity':'2','unit':'LT',
                              'unit_cost':BIRIM_MALIYET,'dose':'25','dose_unit':'ml/da'})
        assert g.status_code == 201, g.text

    # GİRDİSİZ faaliyet — çizelgeden düşmemeli.
    toprak = client.post('/api/field-activities', headers=h,
                         json={'season_id':s1['id'],'activity_type':'TILLAGE',
                               'performed_at':'2026-03-05T09:00:00+03:00'})
    assert toprak.status_code == 201, toprak.text

    # GECE YARISI SINIRI: UTC'de 2026-06-01, İstanbul'da 2026-06-02.
    gece = client.post('/api/field-activities', headers=h,
                       json={'season_id':s1['id'],'activity_type':'IRRIGATION',
                             'performed_at':'2026-06-01T22:30:00+00:00'})
    assert gece.status_code == 201, gece.text

    for tarih, miktar in (('2026-07-10', '12000'), ('2026-07-20', '8000')):
        hv = client.post('/api/field-harvests', headers=h,
                         json={'season_id':s1['id'],'harvested_on':tarih,
                               'quantity':miktar,'unit':'KG','notes':ZEHIR})
        assert hv.status_code == 201, hv.text

    # Başka parselin sezonu — parsel süzgecinde SIZMAMALI.
    client.post('/api/field-activities', headers=h,
                json={'season_id':s2['id'],'activity_type':'IRRIGATION',
                      'performed_at':'2026-06-01T09:00:00+03:00'})

    # --- 2. SÜZGEÇ ZORUNLU -------------------------------------------------
    bos = client.get('/api/exports/producer-logbook', headers=h)
    assert bos.status_code == 400, bos.text
    yalniz_tarih = client.get('/api/exports/producer-logbook',
                              headers=h, params={'date_from':'2026-01-01'})
    assert yalniz_tarih.status_code == 400, yalniz_tarih.text
    ters = client.get('/api/exports/producer-logbook', headers=h,
                      params={'farm_id':ciftlik['id'],
                              'date_from':'2026-12-31','date_to':'2026-01-01'})
    assert ters.status_code == 400, ters.text

    # --- ÖNİZLEME ----------------------------------------------------------
    r = client.get('/api/exports/producer-logbook', headers=h,
                   params={'parcel_id':p1['id']})
    assert r.status_code == 200, r.text
    veri = r.json()
    assert veri['activity_headers'] == list(FAALIYET_BASLIKLARI)
    assert veri['harvest_headers'] == list(HASAT_BASLIKLARI)
    assert 'Resmî bir form değildir' in veri['note']

    faal = veri['activity_rows']
    hasat = veri['harvest_rows']
    basliga = {ad: i for i, ad in enumerate(FAALIYET_BASLIKLARI)}

    # --- 3. KARTEZYEN ÇARPIM YOK -------------------------------------------
    # 2 girdili ilaçlama -> 2 satır; girdisiz iki faaliyet -> 1'er satır.
    assert len(faal) == 4, faal
    # ÜÇ faaliyet olmasına rağmen hasat satırı 2 — çarpmadı.
    assert len(hasat) == 2, hasat
    assert veri['row_count'] == 6, veri['row_count']

    # --- 4. GİRDİSİZ FAALİYET ÇİZELGEDE ------------------------------------
    turler = [s[basliga['Faaliyet']] for s in faal]
    assert turler.count('Toprak işleme') == 1, turler
    assert turler.count('Sulama') == 1, turler
    assert turler.count('İlaçlama') == 2, turler
    (toprak_satiri,) = [s for s in faal if s[basliga['Faaliyet']] == 'Toprak işleme']
    assert toprak_satiri[basliga['Girdi Adı']] == '', toprak_satiri

    # --- İSTANBUL GÜNÜ -----------------------------------------------------
    (gece_satiri,) = [s for s in faal if s[basliga['Faaliyet']] == 'Sulama']
    # UTC 2026-06-01T22:30 -> İstanbul 2026-06-02. UTC gününe düşülseydi 01.06.
    assert gece_satiri[basliga['Uygulama Tarihi']] == '02.06.2026', gece_satiri
    assert gece_satiri[basliga['Uygulama Saati']] == '01:30', gece_satiri

    # Kimlik alanları gerçekten dolu.
    # Kodlar uçta BÜYÜK harfe çevriliyor; çizelge onları OLDUĞU GİBİ taşıyor.
    assert toprak_satiri[basliga['Çiftlik Kodu']].lower() == 'd1', toprak_satiri
    assert toprak_satiri[basliga['Çiftlik']] == 'Defter Çiftlik'
    assert toprak_satiri[basliga['Parsel No']] == '12'
    assert toprak_satiri[basliga['Ada']] == '34'
    assert toprak_satiri[basliga['Ürün']] == 'Buğday'
    assert toprak_satiri[basliga['Çeşit']] == 'Bezostaja'
    (ilac_satiri,) = [s for s in faal if s[basliga['Girdi Adı']] == 'Normal İlaç']
    assert ilac_satiri[basliga['Hasat Bekleme Süresi (gün)']] == 21, ilac_satiri
    assert ilac_satiri[basliga['Tekrar Giriş Süresi (gün)']] == 2, ilac_satiri
    # ÖLÇEK DİYALEKTE BAĞLI: SQLite NUMERIC(18,4)'ü '25', PostgreSQL
    # Decimal('25.0000') olarak veriyor. Metni pinlemek testi bir sürücüye
    # bağlardı; DEĞER karşılaştırılıyor.
    assert Decimal(ilac_satiri[basliga['Doz']]) == Decimal('25'), ilac_satiri
    assert Decimal(ilac_satiri[basliga['Miktar']]) == Decimal('2'), ilac_satiri

    # --- 1a. FORMÜL ENJEKSİYONU: HAM SATIR hâlâ ham ------------------------
    # Üretici arındırmıyor; arındırma xlsx yazımında `_safe_excel_cell` ile
    # oluyor. Bu ayrımı ÇİVİLİYORUZ ki JSON tüketicisi veriyi olduğu gibi
    # görsün ve xlsx yolunun kapısı ayrıca ölçülebilir kalsın.
    assert any(s[basliga['Girdi Adı']] == ZEHIR for s in faal), faal

    # --- SIZINTI: başka parsel yok -----------------------------------------
    assert all(s[basliga['Parsel Kodu']].lower() == 'dp1' for s in faal), faal

    # --- TARİH ARALIĞI KAPSAM SÜZGECİYLE BİRLİKTE --------------------------
    dar = client.get('/api/exports/producer-logbook', headers=h,
                     params={'parcel_id':p1['id'],
                             'date_from':'2026-05-01','date_to':'2026-05-31'}).json()
    assert len(dar['activity_rows']) == 2, dar['activity_rows']
    assert dar['harvest_rows'] == [], dar['harvest_rows']

    # --- 5. PARA SÜTUNU YOK ------------------------------------------------
    tum_basliklar = ' '.join(FAALIYET_BASLIKLARI + HASAT_BASLIKLARI).lower()
    for yasak in ('maliyet', 'tutar', 'fiyat', 'gelir'):
        assert yasak not in tum_basliklar, yasak
    # Değer düzeyinde de: girilen birim maliyet hiçbir hücrede geçmiyor.
    assert not any(BIRIM_MALIYET in str(deger) for s in faal for deger in s), faal

    # --- XLSX --------------------------------------------------------------
    x = client.get('/api/exports/producer-logbook.xlsx', headers=h,
                   params={'parcel_id':p1['id']})
    assert x.status_code == 200, x.text
    assert 'uygulama-kayit-cizelgesi.xlsx' in x.headers['content-disposition']
    wb = load_workbook(BytesIO(x.content))
    assert wb.sheetnames == ['Uygulamalar', 'Hasatlar'], wb.sheetnames

    sh = wb['Uygulamalar']
    # 1. satır not, 2. boş, 3. başlık, sonra satırlar.
    assert 'Resmî bir form değildir' in str(sh.cell(row=1, column=1).value)
    assert [c.value for c in sh[3]] == list(FAALIYET_BASLIKLARI)
    assert sh.max_row == 3 + len(faal), sh.max_row

    # --- 1b. FORMÜL ENJEKSİYONU: HÜCRE arındırılmış ------------------------
    girdi_sutunu = list(FAALIYET_BASLIKLARI).index('Girdi Adı') + 1
    not_sutunu = list(FAALIYET_BASLIKLARI).index('Not') + 1
    hucreler = [sh.cell(row=r, column=girdi_sutunu).value
                for r in range(4, sh.max_row + 1)]
    assert ZEHIR not in hucreler, hucreler
    assert "'" + ZEHIR in hucreler, hucreler
    notlar = [sh.cell(row=r, column=not_sutunu).value
              for r in range(4, sh.max_row + 1)]
    assert "'" + ZEHIR in notlar, notlar
    # Hiçbir hücre çıplak formül karakteriyle BAŞLAMIYOR.
    for satir in sh.iter_rows(min_row=4):
        for hucre in satir:
            if isinstance(hucre.value, str) and hucre.value.strip():
                assert hucre.value.lstrip()[:1] not in {'=', '+', '-', '@'}, hucre.value

    hsh = wb['Hasatlar']
    assert [c.value for c in hsh[3]] == list(HASAT_BASLIKLARI)
    hasat_not = list(HASAT_BASLIKLARI).index('Not') + 1
    assert "'" + ZEHIR == hsh.cell(row=4, column=hasat_not).value

    # --- 6. ÇAPRAZ KİRACI --------------------------------------------------
    b = client.post('/api/companies', headers=h, json={'name':'Defter B'}).json()
    hb = dict(h, **{'X-Company-ID': str(b['id'])})
    capraz = client.get('/api/exports/producer-logbook', headers=hb,
                        params={'parcel_id':p1['id']})
    # 404 DEĞİL 200+BOŞ: uç bir KAYNAĞI değil bir SÜZGECİ okuyor; başka
    # firmanın parsel kimliği bu firmada hiçbir sezona çözülmez. 404 döndürmek,
    # "o kimlik başka yerde VAR" bilgisini sızdırırdı.
    assert capraz.status_code == 200, capraz.text
    assert capraz.json()['activity_rows'] == [], capraz.json()
    assert capraz.json()['harvest_rows'] == [], capraz.json()
    assert capraz.json()['row_count'] == 0
    cx = client.get('/api/exports/producer-logbook.xlsx', headers=hb,
                    params={'farm_id':ciftlik['id']})
    assert cx.status_code == 200, cx.text
    cwb = load_workbook(BytesIO(cx.content))
    assert cwb['Uygulamalar'].max_row == 3, cwb['Uygulamalar'].max_row

    print('URETICI KAYIT DEFTERI CIZELGESI TAMAM')
'''
