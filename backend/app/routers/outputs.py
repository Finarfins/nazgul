from __future__ import annotations

import re
from datetime import date
from io import BytesIO
from xml.sax.saxutils import escape

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field
from qrcode import QRCode
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import text
from sqlalchemy.orm import Session

from ..auth import has_permission
from ..db import get_db
from ..labels import (
    DEFAULT_HEIGHT_MM,
    DEFAULT_WIDTH_MM,
    MAX_COPIES,
    MAX_LABELS,
    MAX_PRODUCTS,
    MAX_SIDE_MM,
    MIN_SIDE_MM,
    LabelFormat,
    LabelOptions,
    PriceVat,
    render_labels,
)
from ..money import money
from ..statement import build_statement, config as statement_config
from ..tenancy import company_id
from ..uretici_kayit_defteri import DefterHatasi, defter_verisi
from ..pdf_fonts import PDF_FONT, PDF_FONT_BOLD, register_pdf_fonts

register_pdf_fonts()

router = APIRouter(tags=["outputs"])
_FORMULA_PREFIX = re.compile(r"^[\t\r\n ]*[=+\-@]")
_SAFE_FILENAME = re.compile(r"[^A-Za-z0-9._-]+")


def _safe_excel_cell(value):
    """Prevent user-controlled values from being interpreted as formulas."""
    if isinstance(value, str) and _FORMULA_PREFIX.match(value):
        return "'" + value
    return value


def _safe_filename(value: object, fallback: str) -> str:
    """Return a bounded ASCII filename fragment safe for response headers."""
    cleaned = _SAFE_FILENAME.sub("-", str(value or "").strip()).strip(".-")
    return cleaned[:100] or fallback


def _require_permission(request: Request, permission: str) -> None:
    user = getattr(request.state, "user", None)
    role = str(user.get("role") or "") if isinstance(user, dict) else ""
    if not has_permission(role, permission):
        raise HTTPException(403, "Bu çıktıyı almak için yetkiniz yok")


def _kind(kind: str):
    if kind == "orders":
        return {
            "head": "orders",
            "items": "order_items",
            "entity": "customers",
            "entity_id": "customer_id",
            "date": "order_date",
            "item_fk": "order_id",
            "title": "SATIŞ BELGESİ",
            "entity_label": "Müşteri",
            "permission": "sales",
        }
    if kind == "purchases":
        return {
            "head": "purchases",
            "items": "purchase_items",
            "entity": "suppliers",
            "entity_id": "supplier_id",
            "date": "purchase_date",
            "item_fk": "purchase_id",
            "title": "ALIŞ BELGESİ",
            "entity_label": "Tedarikçi",
            "permission": "purchases",
        }
    raise HTTPException(404, "Belge türü bulunamadı")


def _document(db: Session, cid: int, kind: str, document_id: int):
    config = _kind(kind)
    head = db.execute(
        text(
            f"""SELECT h.*, e.name entity_name,
            h.{config['date']} transaction_date, w.name warehouse_name
            FROM {config['head']} h
            JOIN {config['entity']} e
              ON e.id=h.{config['entity_id']} AND e.company_id=h.company_id
            LEFT JOIN warehouses w
              ON w.id=h.warehouse_id AND w.company_id=h.company_id
            WHERE h.id=:id AND h.company_id=:cid"""
        ),
        {"id": document_id, "cid": cid},
    ).mappings().first()
    if not head:
        raise HTTPException(404, "Belge bulunamadı")

    lines = db.execute(
        text(
            f"""SELECT product_name, quantity, unit_price, vat_rate,
            COALESCE(discount_percent,0) discount_percent,
            line_subtotal, line_vat, line_total
            FROM {config['items']}
            WHERE {config['item_fk']}=:id
            ORDER BY id"""
        ),
        {"id": document_id},
    ).mappings().all()

    reference_type = "order" if kind == "orders" else "purchase"
    payments = db.execute(
        text(
            """SELECT payment_method, amount, payment_date
            FROM payments
            WHERE company_id=:cid AND reference_type=:rt AND reference_id=:rid
            ORDER BY id"""
        ),
        {"cid": cid, "rt": reference_type, "rid": document_id},
    ).mappings().all()

    result = dict(head)
    result["payments"] = [dict(row) for row in payments]
    result["paid_total"] = sum(
        (money(row["amount"]) for row in payments), money(0)
    )
    result["remaining_total"] = max(
        money(0), money(result.get("final_total")) - result["paid_total"]
    )
    return config, result, [dict(row) for row in lines]


def _money(value):
    return (
        f"{money(value):,.2f} TL"
        .replace(",", "X")
        .replace(".", ",")
        .replace("X", ".")
    )


@router.get("/documents/{kind}/{document_id}/pdf")
def document_pdf(
    kind: str,
    document_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    config = _kind(kind)
    _require_permission(request, config["permission"])
    config, head, lines = _document(db, company_id(request), kind, document_id)

    out = BytesIO()
    document = SimpleDocTemplate(
        out,
        pagesize=A4,
        rightMargin=18 * mm,
        leftMargin=18 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )
    styles = getSampleStyleSheet()
    for style in styles.byName.values():
        style.fontName = PDF_FONT
    story = [Paragraph(config["title"], styles["Title"]), Spacer(1, 5 * mm)]

    meta = [
        [
            "Belge No",
            head.get("document_no") or f"#{head['id']}",
            "Tarih",
            head.get("transaction_date") or "",
        ],
        [
            config["entity_label"],
            head.get("entity_name") or "",
            "Depo",
            head.get("warehouse_name") or "Merkez Depo",
        ],
        [
            "Durum",
            head.get("status") or "completed",
            "Ödeme",
            head.get("payment_method") or "credit",
        ],
    ]
    meta_table = Table(meta, colWidths=[28 * mm, 62 * mm, 28 * mm, 55 * mm])
    meta_table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.35, colors.grey),
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#eef2f6")),
                ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#eef2f6")),
                ("FONTNAME", (0, 0), (-1, -1), PDF_FONT),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("PADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story += [meta_table, Spacer(1, 6 * mm)]

    item_data = [["Ürün", "Miktar", "Birim Fiyat", "KDV", "İsk.", "Toplam"]]
    for row in lines:
        item_data.append(
            [
                row["product_name"],
                f"{row['quantity']:g}",
                _money(row["unit_price"]),
                f"%{row['vat_rate']}",
                f"%{row['discount_percent']:g}",
                _money(row["line_total"]),
            ]
        )
    item_data += [
        ["", "", "", "", "Ara Toplam", _money(head.get("subtotal"))],
        ["", "", "", "", "KDV", _money(head.get("vat_total"))],
        ["", "", "", "", "Genel Toplam", _money(head.get("final_total"))],
    ]
    item_table = Table(
        item_data,
        repeatRows=1,
        colWidths=[68 * mm, 18 * mm, 27 * mm, 16 * mm, 16 * mm, 30 * mm],
    )
    item_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#17324d")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), PDF_FONT_BOLD),
                ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                ("GRID", (0, 0), (-1, -4), 0.3, colors.grey),
                ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
                ("FONTNAME", (4, -3), (5, -1), PDF_FONT_BOLD),
                ("LINEABOVE", (4, -3), (5, -3), 0.7, colors.black),
                ("PADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    story.append(item_table)

    payment_labels = {
        "cash": "Nakit",
        "card": "Kredi Kartı / POS",
        "bank_transfer": "Havale / EFT",
        "check": "Çek",
        "promissory_note": "Senet",
        "credit": "Veresiye",
    }
    payments = head.get("payments") or []
    story += [Spacer(1, 5 * mm), Paragraph("<b>Ödeme Özeti</b>", styles["Heading3"])]
    payment_data = [["Yöntem", "Tutar"]]
    for payment in payments:
        payment_data.append(
            [
                payment_labels.get(
                    payment.get("payment_method"),
                    payment.get("payment_method") or "Diğer",
                ),
                _money(payment.get("amount")),
            ]
        )
    if not payments:
        payment_data.append(["Veresiye / ödeme yok", _money(0)])
    payment_data += [
        ["Ödenen Toplam", _money(head.get("paid_total"))],
        ["Kalan", _money(head.get("remaining_total"))],
    ]
    payment_table = Table(payment_data, colWidths=[90 * mm, 45 * mm], hAlign="RIGHT")
    payment_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef2f6")),
                ("FONTNAME", (0, 0), (-1, 0), PDF_FONT_BOLD),
                ("FONTNAME", (0, -2), (-1, -1), PDF_FONT_BOLD),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
                ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
                ("PADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    story.append(payment_table)

    if head.get("note"):
        safe_note = escape(str(head["note"]))
        story += [
            Spacer(1, 5 * mm),
            Paragraph(f"<b>Not:</b> {safe_note}", styles["BodyText"]),
        ]

    document.build(story)
    out.seek(0)
    number = _safe_filename(
        head.get("document_no") or document_id,
        str(document_id),
    )
    filename = f"{kind}-{number}.pdf"
    return StreamingResponse(
        out,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{filename}"'},
    )


@router.get("/documents/{kind}/{document_id}/xlsx")
def document_xlsx(
    kind: str,
    document_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    config = _kind(kind)
    _require_permission(request, config["permission"])
    config, head, lines = _document(db, company_id(request), kind, document_id)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Belge"
    sheet.append([config["title"]])
    sheet["A1"].font = Font(size=16, bold=True)
    sheet.append([])

    metadata = [
        ["Belge No", head.get("document_no") or head["id"]],
        ["Tarih", head.get("transaction_date")],
        [config["entity_label"], head.get("entity_name")],
        ["Depo", head.get("warehouse_name") or "Merkez Depo"],
        ["Durum", head.get("status")],
        ["Ödeme", head.get("payment_method")],
    ]
    for row in metadata:
        sheet.append([_safe_excel_cell(value) for value in row])

    sheet.append([])
    sheet.append(
        [
            "Ürün",
            "Miktar",
            "Birim Fiyat",
            "KDV %",
            "İskonto %",
            "Ara Toplam",
            "KDV",
            "Toplam",
        ]
    )
    header_row = sheet.max_row
    for cell in sheet[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="17324D")

    for row in lines:
        sheet.append(
            [
                _safe_excel_cell(value)
                for value in [
                    row["product_name"],
                    row["quantity"],
                    row["unit_price"],
                    row["vat_rate"],
                    row["discount_percent"],
                    row["line_subtotal"],
                    row["line_vat"],
                    row["line_total"],
                ]
            ]
        )

    sheet.append([])
    sheet.append(["", "", "", "", "", "", "Genel Toplam", head.get("final_total")])
    widths = {
        "A": 38,
        "B": 14,
        "C": 16,
        "D": 12,
        "E": 14,
        "F": 16,
        "G": 14,
        "H": 16,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center")

    out = BytesIO()
    workbook.save(out)
    out.seek(0)
    number = _safe_filename(
        head.get("document_no") or document_id,
        str(document_id),
    )
    filename = f"{kind}-{number}.xlsx"
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


_PRODUCT_LABEL_COLUMNS = (
    "id,name,product_code,barcode,oem_number,brand,unit,sale_price,vat_rate"
)


def _product(db: Session, cid: int, product_id: int):
    row = db.execute(
        text(
            f"SELECT {_PRODUCT_LABEL_COLUMNS} FROM products "
            "WHERE id=:id AND company_id=:cid"
        ),
        {"id": product_id, "cid": cid},
    ).mappings().first()
    if not row:
        raise HTTPException(404, "Ürün bulunamadı")
    return dict(row)


def _label_products(db: Session, cid: int, product_ids: list[int]) -> list[dict]:
    """Toplu etiket için ürünleri SIRASI KORUNARAK getir.

    Sorgu yalnızca aktif firmanın ürünlerini görür; istenen kimliklerden biri
    bu firmaya ait değilse (ya da hiç yoksa) tekil uçla aynı davranış gösterilir
    ve 404 döner — başka firmanın ürününe etiket üretilmez.
    """
    placeholders = {f"p{index}": value for index, value in enumerate(product_ids)}
    names = ",".join(f":{key}" for key in placeholders)
    rows = db.execute(
        text(
            f"SELECT {_PRODUCT_LABEL_COLUMNS} FROM products "
            f"WHERE company_id=:cid AND id IN ({names})"
        ),
        {"cid": cid, **placeholders},
    ).mappings().all()
    found = {int(row["id"]): dict(row) for row in rows}
    missing = [value for value in product_ids if value not in found]
    if missing:
        raise HTTPException(404, "Ürün bulunamadı")
    return [found[value] for value in product_ids]


def _label_options(
    label_format: LabelFormat,
    show_price: bool,
    price_vat: PriceVat,
    copies: int,
    width_mm: int,
    height_mm: int,
) -> LabelOptions:
    return LabelOptions(
        label_format=label_format,
        show_price=show_price,
        price_vat=price_vat,
        copies=copies,
        width_mm=width_mm,
        height_mm=height_mm,
    )


def _label_response(
    products: list[dict],
    options: LabelOptions,
    filename: str,
) -> StreamingResponse:
    """Etiket PDF'ini üret; toplam etiket sayısını sınırla."""
    total = len(products) * options.copies
    if total > MAX_LABELS:
        raise HTTPException(
            400,
            f"Tek seferde en fazla {MAX_LABELS} etiket üretilebilir "
            f"(istenen: {total}). Ürün sayısını veya kopya adedini azaltın.",
        )
    payload = render_labels(products, options)
    return StreamingResponse(
        BytesIO(payload),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{filename}"'},
    )


@router.get("/products/{product_id}/qr.png")
def product_qr(
    product_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    _require_permission(request, "stock")
    product = _product(db, company_id(request), product_id)
    payload = (
        f"YHP:PRODUCT:{product['id']}:"
        f"{product.get('product_code') or ''}:"
        f"{product.get('barcode') or ''}"
    )
    qr = QRCode(box_size=8, border=2)
    qr.add_data(payload)
    qr.make(fit=True)
    image = qr.make_image(fill_color="black", back_color="white")
    out = BytesIO()
    image.save(out, format="PNG")
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="image/png",
        headers={"Content-Disposition": f'inline; filename="urun-{product_id}-qr.png"'},
    )


class ProductLabelRequest(BaseModel):
    """Toplu etiket gövdesi: ürün listesi + tekil uçla aynı seçenekler."""

    product_ids: list[int] = Field(min_length=1, max_length=MAX_PRODUCTS)
    format: LabelFormat = "thermal"
    show_price: bool = True
    price_vat: PriceVat = "incl"
    copies: int = Field(default=1, ge=1, le=MAX_COPIES)
    width_mm: int = Field(default=DEFAULT_WIDTH_MM, ge=MIN_SIDE_MM, le=MAX_SIDE_MM)
    height_mm: int = Field(default=DEFAULT_HEIGHT_MM, ge=MIN_SIDE_MM, le=MAX_SIDE_MM)


@router.get("/products/{product_id}/label.pdf")
def product_label(
    product_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Eski QR etiketi — ÇIKTI SÖZLEŞMESİ DONDURULMUŞTUR.

    Bu uç sahadaki yazıcı/şablon ayarlarına bağlıdır: QR sembolü, 70x40 mm sayfa
    ve her zaman görünen satış fiyatı. Parametre EKLENMEZ ve varsayılan çıktı
    DEĞİŞTİRİLMEZ; Code128 barkod etiketi ayrı bir uçtur (bkz.
    ``product_barcode_label``). ``test_product_labels.py`` içindeki kilitleyici
    regresyon bu sözleşmeyi çivilemektedir.
    """
    _require_permission(request, "stock")
    product = _product(db, company_id(request), product_id)
    out = BytesIO()
    width, height = 70 * mm, 40 * mm
    pdf = canvas.Canvas(out, pagesize=(width, height))
    pdf.setFont(PDF_FONT_BOLD, 10)
    pdf.drawString(5 * mm, 32 * mm, (product["name"] or "")[:38])
    pdf.setFont(PDF_FONT, 8)
    pdf.drawString(5 * mm, 27 * mm, f"Kod: {product.get('product_code') or '-'}")
    pdf.drawString(5 * mm, 23 * mm, f"Barkod: {product.get('barcode') or '-'}")
    pdf.setFont(PDF_FONT_BOLD, 12)
    pdf.drawString(5 * mm, 15 * mm, _money(product.get("sale_price")))

    payload = (
        f"YHP:PRODUCT:{product['id']}:"
        f"{product.get('product_code') or ''}:"
        f"{product.get('barcode') or ''}"
    )
    qr = QRCode(box_size=4, border=1)
    qr.add_data(payload)
    qr.make(fit=True)
    image = qr.make_image(fill_color="black", back_color="white")
    temporary = BytesIO()
    image.save(temporary, format="PNG")
    temporary.seek(0)

    from reportlab.lib.utils import ImageReader

    pdf.drawImage(
        ImageReader(temporary),
        48 * mm,
        4 * mm,
        width=18 * mm,
        height=18 * mm,
        mask="auto",
    )
    pdf.showPage()
    pdf.save()
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'inline; filename="urun-{product_id}-etiket.pdf"'
        },
    )


@router.get("/products/{product_id}/barcode-label.pdf")
def product_barcode_label(
    product_id: int,
    request: Request,
    format: LabelFormat = Query("thermal"),
    show_price: bool = Query(True),
    price_vat: PriceVat = Query("incl"),
    copies: int = Query(1, ge=1, le=MAX_COPIES),
    width_mm: int = Query(DEFAULT_WIDTH_MM, ge=MIN_SIDE_MM, le=MAX_SIDE_MM),
    height_mm: int = Query(DEFAULT_HEIGHT_MM, ge=MIN_SIDE_MM, le=MAX_SIDE_MM),
    db: Session = Depends(get_db),
):
    """Tek ürün için Code128 barkod etiketi (termal veya A4 ızgara).

    Eski QR etiketinin yerini ALMAZ: ``label.pdf`` dokunulmadan durur, barkod
    etiketi açıkça bu uçtan istenir.
    """
    _require_permission(request, "stock")
    product = _product(db, company_id(request), product_id)
    options = _label_options(
        format, show_price, price_vat, copies, width_mm, height_mm
    )
    return _label_response(
        [product], options, f"urun-{product_id}-barkod-etiket.pdf"
    )


@router.post("/products/labels.pdf")
def product_labels(
    payload: ProductLabelRequest,
    request: Request,
    db: Session = Depends(get_db),
):
    """Birden çok ürün için tek PDF'te toplu barkod etiketi."""
    _require_permission(request, "stock")
    products = _label_products(db, company_id(request), payload.product_ids)
    options = _label_options(
        payload.format,
        payload.show_price,
        payload.price_vat,
        payload.copies,
        payload.width_mm,
        payload.height_mm,
    )
    return _label_response(products, options, "urun-etiketleri.pdf")


@router.get("/exports/products.xlsx")
def products_xlsx(request: Request, db: Session = Depends(get_db)):
    _require_permission(request, "stock")
    cid = company_id(request)
    rows = db.execute(
        text(
            "SELECT id,name,product_code,barcode,category,purchase_price,"
            "sale_price,stock,unit,vat_rate "
            "FROM products WHERE company_id=:cid ORDER BY name"
        ),
        {"cid": cid},
    ).mappings().all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Ürünler"
    headers = [
        "ID",
        "Ürün",
        "Kod",
        "Barkod",
        "Kategori",
        "Alış Fiyatı",
        "Satış Fiyatı",
        "Stok",
        "Birim",
        "KDV %",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="17324D")
    for row in rows:
        sheet.append([_safe_excel_cell(value) for value in row.values()])
    for column, width in zip(
        "ABCDEFGHIJ",
        [8, 38, 18, 18, 20, 16, 16, 12, 10, 10],
    ):
        sheet.column_dimensions[column].width = width

    out = BytesIO()
    workbook.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="urunler.xlsx"'},
    )


# --- Uygulama Kayıt Çizelgesi ------------------------------------------------
#
# YOL ADI BİR GÜVENLİK KARARIDIR. Bu uç `/api/field-…` altına KONMADI:
# `auth.py`de `_FARM_PATH_PREFIXES` listesinde OLMAYAN her `/api/field-` yolu
# sessizce `field_service` iznine düşüyor ve bu tuzağa iki kez düşüldü. Uç
# `/api/exports/` altında duruyor, dolayısıyla middleware `read` çözüyor ve
# yetki KAPISI handler'da açıkça çağrılıyor (`_require_permission`,
# `farm.view`) — tıpkı `/api/exports/products.xlsx` gibi.
#
# Bunun ölçülebilir sonucu: uçlar `GUARDED_READ_OPERATIONS`a düşer, yani rol
# değerine göre REDDEDİLEBİLİR sayılır ve `EXPECTED_UNDENIABLE` popülasyonunu
# BÜYÜTMEZ. Yolu `_FARM_PATH_PREFIXES`e eklemek yetkiyi middleware'e taşır ve
# handler kapısını gereksiz kılardı: aynı erişim, daha zayıf denetim duruşu.


def _logbook_filters(
    farm_id: int | None,
    parcel_id: int | None,
    season_id: int | None,
    season_year: int | None,
    date_from: date | None,
    date_to: date | None,
) -> dict[str, object]:
    return {
        "farm_id": farm_id,
        "parcel_id": parcel_id,
        "season_id": season_id,
        "season_year": season_year,
        "date_from": date_from,
        "date_to": date_to,
    }


def _logbook_data(request: Request, db: Session, filters: dict[str, object]):
    _require_permission(request, "farm.view")
    try:
        return defter_verisi(db, company_id(request), filters)
    except DefterHatasi as hata:
        # Çağıranın DÜZELTEBİLECEĞİ hata: süzgeç eksik, tarih ters ya da sonuç
        # üst sınırı aşıyor. 500 değil 400; metin kullanıcıya gösterilecek.
        raise HTTPException(400, str(hata)) from hata


@router.get("/exports/producer-logbook")
def producer_logbook(
    request: Request,
    db: Session = Depends(get_db),
    farm_id: int | None = Query(None),
    parcel_id: int | None = Query(None),
    season_id: int | None = Query(None),
    season_year: int | None = Query(None),
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
):
    """Çizelgenin JSON hâli — indirmeden önce kapsamı görmek için.

    xlsx ucuyla AYNI üreticiyi çağırır; ekranın "kaç satır çıkacak" sorusunu
    dosyayı indirmeden yanıtlaması için var. İki uç ayrı üretici kullansaydı
    önizleme ile dosya sessizce ayrışabilirdi.
    """
    return _logbook_data(
        request,
        db,
        _logbook_filters(farm_id, parcel_id, season_id, season_year, date_from, date_to),
    )


@router.get("/exports/producer-logbook.xlsx")
def producer_logbook_xlsx(
    request: Request,
    db: Session = Depends(get_db),
    farm_id: int | None = Query(None),
    parcel_id: int | None = Query(None),
    season_id: int | None = Query(None),
    season_year: int | None = Query(None),
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
):
    """İki sayfalı çizelge: uygulamalar ve hasatlar."""
    data = _logbook_data(
        request,
        db,
        _logbook_filters(farm_id, parcel_id, season_id, season_year, date_from, date_to),
    )

    workbook = Workbook()
    activity_sheet = workbook.active
    activity_sheet.title = "Uygulamalar"
    harvest_sheet = workbook.create_sheet("Hasatlar")

    for sheet, headers, rows, widths in (
        (
            activity_sheet,
            data["activity_headers"],
            data["activity_rows"],
            [14, 26, 14, 26, 10, 10, 18, 14, 14, 14, 10, 18, 16, 14, 14, 10, 16, 16,
             28, 12, 10, 10, 12, 16, 18, 20, 20, 16, 30, 34],
        ),
        (
            harvest_sheet,
            data["harvest_headers"],
            data["harvest_rows"],
            [14, 26, 14, 26, 10, 10, 10, 18, 16, 14, 12, 10, 16, 14, 10, 34, 30, 34],
        ),
    ):
        # 1. satır çizelgenin NE OLDUĞUNU söyler. Gerekçe yalnız kod yorumunda
        # kalırsa dosyayı açan kişiye HİÇ ulaşmaz.
        sheet.append([data["note"]])
        sheet.cell(row=1, column=1).font = Font(italic=True, color="7A4A00")
        sheet.append([])
        sheet.append(list(headers))
        for cell in sheet[3]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="17324D")
        for row in rows:
            # SERBEST METİN müfettişin açacağı hücreye gidiyor: `input_name`,
            # `notes`, gerekçeler ve `safety_warning` kullanıcı yazımıdır.
            sheet.append([_safe_excel_cell(value) for value in row])
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width

    out = BytesIO()
    workbook.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition":
                'attachment; filename="uygulama-kayit-cizelgesi.xlsx"'
        },
    )


# Ekstre satır tablosunun stili. Modül düzeyinde durur ki hem PDF üreticisi hem
# de regresyon testi AYNI komut listesini kullansın: gövde hücrelerinin Unicode
# fonta sabitlenmesi (ilk FONTNAME komutu) testin koruduğu asıl davranıştır —
# ReportLab'in tablo varsayılanı Helvetica'dır ve o font Türkçe glifleri
# taşımaz (bkz. app/pdf_fonts.py).
STATEMENT_TABLE_STYLE = [
    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#17324d")),
    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
    ("FONTNAME", (0, 0), (-1, -1), PDF_FONT),
    ("FONTNAME", (0, 0), (-1, 0), PDF_FONT_BOLD),
    ("FONTNAME", (0, 1), (-1, 1), PDF_FONT_BOLD),
    ("FONTNAME", (0, -1), (-1, -1), PDF_FONT_BOLD),
    ("FONTSIZE", (0, 0), (-1, -1), 8.5),
    ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
    ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
    ("LINEABOVE", (0, -1), (-1, -1), 0.7, colors.black),
    ("PADDING", (0, 0), (-1, -1), 5),
]


def _date_tr(value: object) -> str:
    """ISO tarihi (``YYYY-MM-DD``) Türkçe biçime çevir (``GG.AA.YYYY``)."""
    text_value = str(value or "")
    if len(text_value) < 10:
        return text_value or "-"
    year, month, day = text_value[:4], text_value[5:7], text_value[8:10]
    return f"{day}.{month}.{year}"


def _statement_pdf(
    entity_type: str,
    entity_id: int,
    request: Request,
    db: Session,
    date_from: date | None,
    date_to: date | None,
) -> StreamingResponse:
    """Cari ekstresini yazdırılabilir PDF olarak üret.

    Belge içeriği JSON ucuyla aynı ``build_statement`` çıktısından gelir, yani
    ekranda görünen ekstre ile elden verilen kâğıt aynı bakiyeyi taşır.
    """
    settings = statement_config(entity_type)
    _require_permission(request, settings["permission"])
    cid = company_id(request)
    statement = build_statement(db, cid, entity_type, entity_id, date_from, date_to)
    company = db.execute(
        text("SELECT id,name,tax_number FROM companies WHERE id=:cid"),
        {"cid": cid},
    ).mappings().first()

    out = BytesIO()
    document = SimpleDocTemplate(
        out,
        pagesize=A4,
        rightMargin=18 * mm,
        leftMargin=18 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
        title=f"Cari Ekstre {statement.entity.name}",
        author="Harman Zamanı",
    )
    styles = getSampleStyleSheet()
    for style in styles.byName.values():
        style.fontName = PDF_FONT

    company_name = escape(str((company or {}).get("name") or "Harman Zamanı"))
    company_tax = escape(str((company or {}).get("tax_number") or "-"))
    story = [
        Paragraph(
            f"<b>{company_name}</b><br/>Vergi No: {company_tax}",
            styles["Normal"],
        ),
        Spacer(1, 4 * mm),
        Paragraph("CARİ HESAP EKSTRESİ", styles["Title"]),
        Spacer(1, 4 * mm),
    ]

    entity = statement.entity
    meta = [
        [settings["entity_label"], entity.name or "", "Dönem",
         f"{_date_tr(statement.date_from)} - {_date_tr(statement.date_to)}"],
        ["VKN / TCKN", entity.tax_number or "-", "Telefon", entity.phone or "-"],
        ["Adres", entity.address or "-", "Yetkili", entity.owner_name or "-"],
    ]
    meta_table = Table(meta, colWidths=[26 * mm, 62 * mm, 26 * mm, 60 * mm])
    meta_table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.35, colors.grey),
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#eef2f6")),
                ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#eef2f6")),
                ("FONTNAME", (0, 0), (-1, -1), PDF_FONT),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("PADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story += [meta_table, Spacer(1, 5 * mm)]

    rows = [["Tarih", "Açıklama", "Belge No", "Borç", "Alacak", "Bakiye"]]
    rows.append(
        [
            _date_tr(statement.date_from),
            "Devir (açılış bakiyesi)",
            "",
            "",
            "",
            _money(statement.opening_balance),
        ]
    )
    for line in statement.lines:
        rows.append(
            [
                _date_tr(line.entry_date),
                line.label,
                line.document_no or "",
                _money(line.debit) if line.debit else "",
                _money(line.credit) if line.credit else "",
                _money(line.balance),
            ]
        )
    rows.append(
        [
            _date_tr(statement.date_to),
            "Dönem toplamı / kapanış bakiyesi",
            "",
            _money(statement.total_debit),
            _money(statement.total_credit),
            _money(statement.closing_balance),
        ]
    )

    line_table = Table(
        rows,
        repeatRows=1,
        colWidths=[20 * mm, 50 * mm, 28 * mm, 24 * mm, 24 * mm, 28 * mm],
    )
    line_table.setStyle(TableStyle(STATEMENT_TABLE_STYLE))
    story.append(line_table)

    if not statement.lines:
        story += [
            Spacer(1, 4 * mm),
            Paragraph("Bu dönemde hareket bulunmuyor.", styles["BodyText"]),
        ]
    if statement.truncated:
        story += [
            Spacer(1, 4 * mm),
            Paragraph(
                f"<b>Not:</b> Dönemdeki hareket sayısı listeleme sınırını "
                f"({len(statement.lines)}) aştığı için satırlar kısaltılmıştır. "
                f"Açılış ve kapanış bakiyeleri dönemin tamamını kapsar; tam "
                f"döküm için daha dar bir tarih aralığı seçin.",
                styles["BodyText"],
            ),
        ]

    document.build(story)
    out.seek(0)
    name = _safe_filename(entity.name, str(entity_id))
    filename = f"ekstre-{name}-{statement.date_from}-{statement.date_to}.pdf"
    return StreamingResponse(
        out,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{filename}"'},
    )


@router.get("/customers/{customer_id}/statement.pdf")
def customer_statement_pdf(
    customer_id: int,
    request: Request,
    date_from: date | None = None,
    date_to: date | None = None,
    db: Session = Depends(get_db),
):
    return _statement_pdf(
        "customer", customer_id, request, db, date_from, date_to
    )


@router.get("/suppliers/{supplier_id}/statement.pdf")
def supplier_statement_pdf(
    supplier_id: int,
    request: Request,
    date_from: date | None = None,
    date_to: date | None = None,
    db: Session = Depends(get_db),
):
    return _statement_pdf(
        "supplier", supplier_id, request, db, date_from, date_to
    )
