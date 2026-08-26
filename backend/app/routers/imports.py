from decimal import Decimal
import csv
from io import BytesIO
from io import StringIO
import logging
import re
from typing import Any
from zipfile import BadZipFile, ZipFile

from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import text
from sqlalchemy.orm import Session

from ..business_time import business_today
from ..company_policies import (
    POLICY_MANAGER_OVERRIDE,
    get_policy_settings,
    negative_stock_allowed,
    override_context,
    record_policy_overrides,
    stock_policy_message,
)
from ..db import get_db
from ..document_engine import SALES_IMPORT_NOTE
from ..schemas import PaymentCreate
from ..money import ZERO, decimal_value, quantity
from ..inventory import adjust_warehouse_stock, default_warehouse, sync_product_stock
from ..tenancy import company_id

router = APIRouter(prefix='/imports', tags=['imports'])
logger = logging.getLogger("yerel_hesap")
IMPORT_FAILED_MESSAGE = "İçe aktarma tamamlanamadı. Lütfen tekrar deneyin."
#: Ürün içe aktarma bir depo gerektirir; eksikse ne yapılacağı söylenmeli.
NO_WAREHOUSE_MESSAGE = (
    "Ürünler bir depoya bağlanır ve bu firmada tanımlı aktif depo yok. "
    "Önce Stok & Ürünler → Depolar ekranından bir depo oluşturun, sonra "
    "yüklemeyi tekrarlayın."
)

CUSTOMER_HEADERS = ['İsim/Ünvan', 'Telefon', 'E-posta', 'Adres', 'Vergi No', 'Açılış Bakiyesi']
PRODUCT_HEADERS = ['Ürün Adı', 'Ürün Kodu', 'Barkod', 'Kategori', 'Alış Fiyatı', 'Satış Fiyatı', 'Stok', 'Birim', 'KDV Oranı']
MAX_EXCEL_UPLOAD_BYTES = 10 * 1024 * 1024
MAX_EXCEL_UNCOMPRESSED_BYTES = 100 * 1024 * 1024
MAX_EXCEL_ROWS = 50_000
UPLOAD_CHUNK_BYTES = 1024 * 1024


async def _read_upload_limited(file: UploadFile) -> bytes:
    chunks: list[bytes] = []
    total = 0
    while True:
        chunk = await file.read(UPLOAD_CHUNK_BYTES)
        if not chunk:
            break
        total += len(chunk)
        if total > MAX_EXCEL_UPLOAD_BYTES:
            raise HTTPException(
                status_code=413,
                detail="Excel dosyası en fazla 10 MB olabilir.",
            )
        chunks.append(chunk)
    return b"".join(chunks)


def _validate_xlsx_archive(raw: bytes) -> None:
    try:
        with ZipFile(BytesIO(raw)) as archive:
            total_uncompressed = sum(item.file_size for item in archive.infolist())
            if total_uncompressed > MAX_EXCEL_UNCOMPRESSED_BYTES:
                raise HTTPException(
                    status_code=413,
                    detail="Excel arşivinin açılmış boyutu güvenli sınırı aşıyor.",
                )
    except BadZipFile as exc:
        raise HTTPException(400, "Geçerli bir .xlsx/.xlsm dosyası değil.") from exc


def _template(headers: list[str], example: list[Any], filename: str):
    wb = Workbook(); ws = wb.active; ws.title = 'Şablon'
    ws.append(headers); ws.append(example)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='17324D')
    for idx, header in enumerate(headers, start=1):
        ws.column_dimensions[chr(64 + idx)].width = max(14, min(34, len(header) + 5))
    out = BytesIO(); wb.save(out); out.seek(0)
    return StreamingResponse(out, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': f'attachment; filename="{filename}"'})


@router.get('/customers/template.xlsx')
def customer_template():
    return _template(CUSTOMER_HEADERS, ['Örnek Müşteri', '05321234567', 'ornek@firma.com', 'İstanbul', '1234567890', 0], 'musteri-sablonu.xlsx')




@router.get('/suppliers/template.xlsx')
def supplier_template():
    return _template(CUSTOMER_HEADERS, ['Örnek Tedarikçi', '05321234567', 'ornek@firma.com', 'İstanbul', '1234567890', 0], 'tedarikci-sablonu.xlsx')


@router.post('/suppliers/excel')
async def import_suppliers(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not (file.filename or '').lower().endswith(('.xlsx', '.xlsm')):
        raise HTTPException(400, 'Yalnızca .xlsx dosyası yükleyin.')
    headers, rows = _read_rows(await _read_upload_limited(file))
    mapping = _map(headers, {
        'name':['İsim/Ünvan','İsim','Tedarikçi','Tedarikçi Adı','name'], 'phone':['Telefon','phone'],
        'email':['E-posta','Email','email'], 'address':['Adres','address'], 'tax_number':['Vergi No','Vergi Numarası','tax_number'],
        'opening_balance':['Açılış Bakiyesi','Bakiye','opening_balance']
    })
    if 'name' not in mapping: raise HTTPException(400, 'İsim/Ünvan sütunu bulunamadı.')
    cid = company_id(request); inserted = updated = 0; errors: list[dict[str, Any]] = []
    try:
        for line_no, row in enumerate(rows, start=2):
            name = str(_cell(row,mapping,'name','') or '').strip()
            if not name:
                errors.append({'row':line_no,'message':'İsim/Ünvan boş.'}); continue
            values={'cid':cid,'name':name,'phone':str(_cell(row,mapping,'phone','') or '').strip() or None,
                    'email':str(_cell(row,mapping,'email','') or '').strip() or None,'address':str(_cell(row,mapping,'address','') or '').strip() or None,
                    'tax_number':str(_cell(row,mapping,'tax_number','') or '').strip() or None,'opening_balance':_num(_cell(row,mapping,'opening_balance',0))}
            existing=db.execute(text('SELECT id FROM suppliers WHERE company_id=:cid AND LOWER(name)=LOWER(:name)'),values).scalar()
            if existing:
                values['id']=existing
                db.execute(text('UPDATE suppliers SET phone=:phone,email=:email,address=:address,tax_number=:tax_number,opening_balance=:opening_balance WHERE id=:id AND company_id=:cid'),values); updated+=1
            else:
                db.execute(text('INSERT INTO suppliers(name,phone,email,address,tax_number,opening_balance,company_id) VALUES(:name,:phone,:email,:address,:tax_number,:opening_balance,:cid)'),values); inserted+=1
        db.commit()
    except Exception as exc:
        db.rollback()
        logger.exception("İçe aktarma beklenmeyen hata")
        raise HTTPException(400, IMPORT_FAILED_MESSAGE) from exc
    return {'inserted':inserted,'updated':updated,'errors':errors,'total_rows':len(rows)}


@router.get('/products/template.xlsx')
def product_template():
    return _template(PRODUCT_HEADERS, ['Örnek Ürün', 'URN-001', '869000000001', 'Genel', 100, 150, 10, 'Adet', 20], 'urun-sablonu.xlsx')


def _norm_header(value: Any) -> str:
    s = str(value or '').strip().lower()
    table = str.maketrans('çğıöşü', 'cgiosu')
    return ''.join(ch for ch in s.translate(table) if ch.isalnum())


def _read_rows(raw: bytes) -> tuple[list[str], list[list[Any]]]:
    _validate_xlsx_archive(raw)
    try:
        wb = load_workbook(BytesIO(raw), data_only=True, read_only=True)
        ws = wb.active
    except Exception as exc:
        raise HTTPException(400, f'Excel dosyası okunamadı: {exc}')
    iterator = ws.iter_rows(values_only=True)
    first = next(iterator, None)
    if first is None:
        raise HTTPException(400, 'Excel dosyası boş.')
    headers = [_norm_header(x) for x in first]
    rows: list[list[Any]] = []
    for row_number, row in enumerate(iterator, start=2):
        if row_number > MAX_EXCEL_ROWS + 1:
            raise HTTPException(
                status_code=413,
                detail=f'Excel dosyası en fazla {MAX_EXCEL_ROWS} veri satırı içerebilir.',
            )
        if any(x not in (None, '') for x in row):
            rows.append(list(row))
    return headers, rows


def _read_csv_rows(raw: bytes) -> tuple[list[str], list[list[Any]]]:
    try:
        content = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            content = raw.decode("cp1254")
        except UnicodeDecodeError as exc:
            raise HTTPException(400, "CSV dosyası UTF-8 veya Windows Türkçe kodlamasında olmalıdır.") from exc
    try:
        dialect = csv.Sniffer().sniff(content[:4096], delimiters=",;\t")
        parsed = csv.reader(StringIO(content), dialect)
        first = next(parsed, None)
        if first is None:
            raise HTTPException(400, "CSV dosyası boş.")
        headers = [_norm_header(value) for value in first]
        rows: list[list[Any]] = []
        for row_number, row in enumerate(parsed, start=2):
            if row_number > MAX_EXCEL_ROWS + 1:
                raise HTTPException(
                    status_code=413,
                    detail=f"CSV dosyası en fazla {MAX_EXCEL_ROWS} veri satırı içerebilir.",
                )
            if any(value.strip() for value in row):
                rows.append(list(row))
        return headers, rows
    except csv.Error as exc:
        raise HTTPException(400, f"CSV dosyası okunamadı: {exc}") from exc


async def _read_tabular_upload(file: UploadFile) -> tuple[list[str], list[list[Any]]]:
    """Shared bounded Excel/CSV upload path for business-data imports."""
    filename = (file.filename or "").lower()
    if not filename.endswith((".xlsx", ".xlsm", ".csv")):
        raise HTTPException(400, "Yalnızca .xlsx, .xlsm veya .csv dosyası yükleyin.")
    raw = await _read_upload_limited(file)
    return _read_csv_rows(raw) if filename.endswith(".csv") else _read_rows(raw)


def _map(headers: list[str], aliases: dict[str, list[str]]) -> dict[str, int]:
    result: dict[str, int] = {}
    for key, names in aliases.items():
        normalized = {_norm_header(x) for x in names}
        for idx, value in enumerate(headers):
            if value in normalized:
                result[key] = idx; break
    return result


def _cell(row: list[Any], mapping: dict[str, int], key: str, default: Any = None):
    idx = mapping.get(key)
    return row[idx] if idx is not None and idx < len(row) else default


def _num(value: Any, default: Decimal | int | str = Decimal("0")) -> Decimal:
    if value in (None, ''):
        return decimal_value(default)
    if not isinstance(value, str):
        return decimal_value(value)

    raw = value.strip().replace('\u00a0', '').replace(' ', '')
    if not raw:
        return decimal_value(default)
    if not re.fullmatch(r'[+-]?[0-9][0-9.,]*', raw):
        raise ValueError(f'Geçersiz sayı: {value!r}')

    sign = ''
    if raw[0] in '+-':
        sign, raw = raw[0], raw[1:]

    dot_count = raw.count('.')
    comma_count = raw.count(',')

    if dot_count and comma_count:
        # The last separator is the decimal separator; all earlier instances
        # are thousands separators. Supports both 1.234,56 and 1,234.56.
        decimal_separator = '.' if raw.rfind('.') > raw.rfind(',') else ','
        thousands_separator = ',' if decimal_separator == '.' else '.'
        normalized = raw.replace(thousands_separator, '')
        normalized = normalized.replace(decimal_separator, '.')
    elif dot_count or comma_count:
        separator = '.' if dot_count else ','
        parts = raw.split(separator)
        if len(parts) > 2:
            if not all(len(group) == 3 for group in parts[1:]):
                raise ValueError(f'Belirsiz sayı biçimi: {value!r}')
            normalized = ''.join(parts)
        else:
            integer, fraction = parts
            # A single separator followed by exactly three digits is ambiguous:
            # "1.234" can mean 1234 in TR formatting or 1.234 in EN formatting.
            # Never guess in an accounting import; require an unambiguous value
            # such as 1234, 1.234,00 or 1,234.00.
            if len(fraction) == 3:
                raise ValueError(
                    f'Belirsiz sayı biçimi: {value!r}. Binlik ve ondalık ayracını açık yazın.'
                )
            normalized = integer + '.' + fraction
    else:
        normalized = raw

    return decimal_value(sign + normalized)


@router.post('/customers/excel')
async def import_customers(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not (file.filename or '').lower().endswith(('.xlsx', '.xlsm')):
        raise HTTPException(400, 'Yalnızca .xlsx dosyası yükleyin.')
    headers, rows = _read_rows(await _read_upload_limited(file))
    mapping = _map(headers, {
        'name':['İsim/Ünvan','İsim','Müşteri','Müşteri Adı','name'], 'phone':['Telefon','phone'],
        'email':['E-posta','Email','email'], 'address':['Adres','address'], 'tax_number':['Vergi No','Vergi Numarası','tax_number'],
        'opening_balance':['Açılış Bakiyesi','Bakiye','opening_balance']
    })
    if 'name' not in mapping: raise HTTPException(400, 'İsim/Ünvan sütunu bulunamadı.')
    cid = company_id(request); inserted = updated = 0; errors: list[dict[str, Any]] = []
    try:
        for line_no, row in enumerate(rows, start=2):
            name = str(_cell(row,mapping,'name','') or '').strip()
            if not name:
                errors.append({'row':line_no,'message':'İsim/Ünvan boş.'}); continue
            values={'cid':cid,'name':name,'phone':str(_cell(row,mapping,'phone','') or '').strip() or None,
                    'email':str(_cell(row,mapping,'email','') or '').strip() or None,'address':str(_cell(row,mapping,'address','') or '').strip() or None,
                    'tax_number':str(_cell(row,mapping,'tax_number','') or '').strip() or None,'opening_balance':_num(_cell(row,mapping,'opening_balance',0))}
            existing=db.execute(text('SELECT id FROM customers WHERE company_id=:cid AND LOWER(name)=LOWER(:name)'),values).scalar()
            if existing:
                values['id']=existing
                db.execute(text('UPDATE customers SET phone=:phone,email=:email,address=:address,tax_number=:tax_number,opening_balance=:opening_balance WHERE id=:id AND company_id=:cid'),values); updated+=1
            else:
                db.execute(text('INSERT INTO customers(name,phone,email,address,tax_number,opening_balance,company_id) VALUES(:name,:phone,:email,:address,:tax_number,:opening_balance,:cid)'),values); inserted+=1
        db.commit()
    except Exception as exc:
        db.rollback()
        logger.exception("İçe aktarma beklenmeyen hata")
        raise HTTPException(400, IMPORT_FAILED_MESSAGE) from exc
    return {'inserted':inserted,'updated':updated,'errors':errors,'total_rows':len(rows)}


@router.post('/products/excel')
async def import_products(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not (file.filename or '').lower().endswith(('.xlsx', '.xlsm')):
        raise HTTPException(400, 'Yalnızca .xlsx dosyası yükleyin.')
    headers, rows = _read_rows(await _read_upload_limited(file))
    mapping = _map(headers, {
        'name':['Ürün Adı','Ürün/Hizmet Adı','Ürün','name'], 'product_code':['Ürün Kodu','Kod','product_code'], 'barcode':['Barkod','barcode'],
        'category':['Kategori','category'], 'purchase_price':['Alış Fiyatı','Alış','purchase_price'], 'sale_price':['Satış Fiyatı','Satış','sale_price'],
        'stock':['Stok','Stok Miktarı','stock'], 'unit':['Birim','unit'], 'vat_rate':['KDV Oranı','KDV','vat_rate']
    })
    if 'name' not in mapping: raise HTTPException(400, 'Ürün Adı sütunu bulunamadı.')
    cid=company_id(request)
    # Ürünler bir depoya bağlanır. Depo yoksa default_warehouse RuntimeError
    # fırlatıyordu ve bu, try bloğunun dışında olduğu için kullanıcıya kuru bir
    # "Aktarım başarısız." olarak düşüyordu: eksiğin depo olduğu anlaşılmıyordu.
    try:
        wid=default_warehouse(db,cid)
    except RuntimeError as exc:
        raise HTTPException(400, NO_WAREHOUSE_MESSAGE) from exc
    inserted=updated=0; errors=[]
    policies=get_policy_settings(db,cid); override=override_context(request); override_policies:set[str]=set()
    try:
        for line_no,row in enumerate(rows,start=2):
            name=str(_cell(row,mapping,'name','') or '').strip()
            if not name:
                errors.append({'row':line_no,'message':'Ürün adı boş.'}); continue
            code=str(_cell(row,mapping,'product_code','') or '').strip() or None
            barcode=str(_cell(row,mapping,'barcode','') or '').strip() or None
            params={'cid':cid,'name':name,'code':code,'barcode':barcode,'category':str(_cell(row,mapping,'category','') or '').strip() or None,
                    'purchase':_num(_cell(row,mapping,'purchase_price',0)),'sale':_num(_cell(row,mapping,'sale_price',0)),
                    'unit':str(_cell(row,mapping,'unit','Adet') or 'Adet').strip(),'vat':int(_num(_cell(row,mapping,'vat_rate',20),20)),'stock':_num(_cell(row,mapping,'stock',0))}
            existing=None
            if code: existing=db.execute(text('SELECT id,stock FROM products WHERE company_id=:cid AND product_code=:code'),params).mappings().first()
            if not existing and barcode: existing=db.execute(text('SELECT id,stock FROM products WHERE company_id=:cid AND barcode=:barcode'),params).mappings().first()
            if not existing: existing=db.execute(text('SELECT id,stock FROM products WHERE company_id=:cid AND LOWER(name)=LOWER(:name)'),params).mappings().first()
            if existing:
                params['id']=existing['id']; old=quantity(existing['stock'])
                db.execute(text('UPDATE products SET name=:name,product_code=:code,barcode=:barcode,category=:category,purchase_price=:purchase,sale_price=:sale,unit=:unit,vat_rate=:vat WHERE id=:id AND company_id=:cid'),params)
                diff=params['stock']-old
                if diff:
                    allow_negative=negative_stock_allowed(policies.negative_stock_policy,override) if diff<0 else False
                    new_stock=adjust_warehouse_stock(db,cid,wid,existing['id'],diff,allow_negative=allow_negative)
                    if new_stock<0 and policies.negative_stock_policy==POLICY_MANAGER_OVERRIDE: override_policies.add('negative_stock')
                    db.execute(text("INSERT INTO stock_movements(product_id,movement_type,quantity,movement_date,reference_type,reference_id,note,company_id,warehouse_id) VALUES(:id,'excel',:q,:movement_date,'import',NULL,'Excel ürün güncelleme',:cid,:wid)"),{'id':existing['id'],'q':diff,'cid':cid,'wid':wid,'movement_date':business_today().isoformat()})
                sync_product_stock(db,cid,existing['id']); updated+=1
            else:
                r=db.execute(text('INSERT INTO products(name,product_code,barcode,category,purchase_price,sale_price,vat_rate,stock,unit,company_id) VALUES(:name,:code,:barcode,:category,:purchase,:sale,:vat,0,:unit,:cid) RETURNING id'),params); pid=int(r.scalar_one())
                if params['stock']:
                    stock_delta=params['stock']; allow_negative=negative_stock_allowed(policies.negative_stock_policy,override) if stock_delta<0 else False
                    new_stock=adjust_warehouse_stock(db,cid,wid,pid,stock_delta,allow_negative=allow_negative)
                    if new_stock<0 and policies.negative_stock_policy==POLICY_MANAGER_OVERRIDE: override_policies.add('negative_stock')
                    db.execute(text("INSERT INTO stock_movements(product_id,movement_type,quantity,movement_date,reference_type,reference_id,note,company_id,warehouse_id) VALUES(:id,'excel',:q,:movement_date,'import',NULL,'Excel ürün açılış stoku',:cid,:wid)"),{'id':pid,'q':params['stock'],'cid':cid,'wid':wid,'movement_date':business_today().isoformat()})
                sync_product_stock(db,cid,pid); inserted+=1
        record_policy_overrides(db,company_id=cid,context=override,policy_names=override_policies,resource_type='imports:products',resource_id=None)
        db.commit()
    except HTTPException:
        db.rollback(); raise
    except ValueError as exc:
        db.rollback()
        if 'yetersiz stok' in str(exc).lower(): raise HTTPException(409,stock_policy_message(policies.negative_stock_policy)) from exc
        raise HTTPException(400,f'Aktarım başarısız: {exc}') from exc
    except Exception as exc:
        db.rollback()
        logger.exception("Ürün içe aktarma beklenmeyen hata", extra={"path": "/api/imports/products"})
        raise HTTPException(400, IMPORT_FAILED_MESSAGE) from exc
    return {'inserted':inserted,'updated':updated,'errors':errors,'total_rows':len(rows),'policy_overrides':sorted(override_policies)}


# --- BizimHesap satış raporu içe aktarma -------------------------------------
# BizimHesap "SATIŞ RAPORU" export'u satır-bazlıdır (her satır bir satış kalemi),
# belge numarası taşımaz ve KDV dahil/hariç bilgisi tutarsızdır. Bu yüzden:
#  * kalemler (müşteri + gün) bazında tek bir satış belgesinde birleştirilir,
#  * birim fiyat, tutarlı olan "Net" (KDV hariç, indirimli) kolonundan türetilir,
#  * belgeler ``draft`` statüsünde girilir; stok zaten güncel haliyle aktarıldığı
#    için tekrar düşmemesi gerekir (STOCK_STATUSES dışı = stok-nötr).
MAX_SALES_IMPORT_ROWS = 20_000
# Aynı dosyanın ikinci kez yüklenmesi tüm satış geçmişini mükerrer oluşturur.
# İçe aktarılan belgeler bu not ile işaretlenir; ikinci yükleme, çağıran açıkça
# ``allow_duplicate=true`` demedikçe 409 ile reddedilir.
def _read_rows_with_header(raw: bytes, must_have: list[str]) -> tuple[list[str], list[list[Any]]]:
    """Read an .xlsx whose header row may follow title/date rows (BizimHesap)."""
    _validate_xlsx_archive(raw)
    try:
        wb = load_workbook(BytesIO(raw), data_only=True, read_only=True)
        ws = wb.active
    except Exception as exc:
        raise HTTPException(400, f'Excel dosyası okunamadı: {exc}')
    wanted = {_norm_header(m) for m in must_have}
    header_idx = None
    all_rows: list[tuple] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        all_rows.append(row)
        if header_idx is None and i < 15:
            if wanted & {_norm_header(c) for c in row}:
                header_idx = i
        if len(all_rows) > MAX_SALES_IMPORT_ROWS + 20:
            raise HTTPException(413, f'Satış raporu en fazla {MAX_SALES_IMPORT_ROWS} satır içerebilir.')
    if header_idx is None:
        raise HTTPException(400, 'Başlık satırı bulunamadı. Beklenen sütunlar: ' + ', '.join(must_have))
    headers = [_norm_header(x) for x in all_rows[header_idx]]
    rows = [list(r) for r in all_rows[header_idx + 1:] if any(x not in (None, '') for x in r)]
    return headers, rows


def _sales_unit_price(*, net: Decimal, gross: Decimal, vat_rate: int, qty: Decimal) -> Decimal:
    """Return the VAT-INCLUSIVE unit price ``_save``/``_totals`` expects.

    ``_totals`` treats ``unit_price`` as VAT-inclusive and extracts VAT from
    within the line total (``line_subtotal = line_total / (1 + vat/100)``).
    BizimHesap's ``Net`` column is VAT-exclusive, so feeding it straight through
    would under-report every document by the VAT amount. The report's ``TOPLAM``
    column is the VAT-inclusive line total and is used when present; otherwise
    VAT is added back onto ``Net``.
    """
    if qty <= 0:
        return decimal_value(0)
    line_gross = gross if gross > 0 else decimal_value(net) * (Decimal(1) + Decimal(vat_rate) / Decimal(100))
    return decimal_value(line_gross) / qty


def _to_iso_date(value: Any) -> str | None:
    if value in (None, ''):
        return None
    if hasattr(value, 'strftime'):
        return value.strftime('%Y-%m-%d')
    s = str(value).strip()
    m = re.fullmatch(r'(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{4})', s)
    if m:
        d, mo, y = m.groups()
        return f'{y}-{int(mo):02d}-{int(d):02d}'
    m = re.match(r'(\d{4})-(\d{2})-(\d{2})', s)
    if m:
        return f'{m.group(1)}-{m.group(2)}-{m.group(3)}'
    return None


@router.post('/sales/excel')
async def import_sales(
    request: Request,
    file: UploadFile = File(...),
    allow_duplicate: bool = False,
    db: Session = Depends(get_db),
):
    """Import a BizimHesap sales report as stock-neutral draft sale documents."""
    if not (file.filename or '').lower().endswith(('.xlsx', '.xlsm')):
        raise HTTPException(400, 'Yalnızca .xlsx dosyası yükleyin.')
    raw = await _read_upload_limited(file)
    headers, rows = _read_rows_with_header(raw, ['Müşteri', 'Ürün'])
    if len(rows) > MAX_SALES_IMPORT_ROWS:
        raise HTTPException(413, f'Satış raporu en fazla {MAX_SALES_IMPORT_ROWS} satır içerebilir.')
    mapping = _map(headers, {
        'customer': ['Müşteri', 'Müşteri Adı', 'Cari'],
        'product': ['Ürün', 'Ürün Adı', 'Ürün/Hizmet Adı'],
        'code': ['Kod', 'Ürün Kodu', 'Stok Kodu'],
        'barcode': ['Barkod', 'Barkodu'],
        'quantity': ['Miktar', 'Adet'],
        'net': ['Net'],
        'gross': ['TOPLAM', 'Toplam'],
        'vat_rate': ['KDV(%)', 'KDV %', 'KDV Oranı'],
        'date': ['Tarih', 'Belge Tarihi'],
    })
    if 'customer' not in mapping:
        raise HTTPException(400, 'Müşteri sütunu bulunamadı.')
    if 'product' not in mapping and 'code' not in mapping:
        raise HTTPException(400, 'Ürün veya Kod sütunu bulunamadı.')
    if 'net' not in mapping:
        raise HTTPException(400, 'Net (KDV hariç tutar) sütunu bulunamadı.')

    cid = company_id(request)
    override = override_context(request)

    if not allow_duplicate:
        already = db.execute(
            text('SELECT COUNT(*) FROM orders WHERE company_id=:cid AND note=:note'),
            {'cid': cid, 'note': SALES_IMPORT_NOTE},
        ).scalar()
        if already:
            raise HTTPException(
                409,
                f'Bu şirkette daha önce {already} adet satış raporu belgesi içe aktarılmış. '
                'Mükerrer kayıt oluşmaması için işlem durduruldu. Yine de devam etmek '
                'istiyorsanız yükleme isteğini allow_duplicate=true ile tekrarlayın.',
            )

    # Tenant-scoped eşleşme tabloları: aynı normalizasyon ürün/cari importuyla tutarlı.
    cust_by_name: dict[str, int] = {}
    for r in db.execute(text('SELECT id,name FROM customers WHERE company_id=:cid'), {'cid': cid}).mappings():
        cust_by_name.setdefault(_norm_header(r['name']), int(r['id']))
    prod_by_code: dict[str, dict] = {}
    prod_by_barcode: dict[str, dict] = {}
    prod_by_name: dict[str, dict] = {}
    for r in db.execute(
        text('SELECT id,name,product_code,barcode,vat_rate FROM products WHERE company_id=:cid'), {'cid': cid}
    ).mappings():
        rec = {'id': int(r['id']), 'vat': r['vat_rate']}
        if r['product_code']:
            prod_by_code.setdefault(_norm_header(r['product_code']), rec)
        if r['barcode']:
            prod_by_barcode.setdefault(_norm_header(r['barcode']), rec)
        prod_by_name.setdefault(_norm_header(r['name']), rec)

    groups: dict[tuple, dict] = {}
    errors: list[dict[str, Any]] = []
    unmatched_customers: set[str] = set()
    unmatched_products: set[str] = set()
    skipped = 0

    for line_no, row in enumerate(rows, start=1):
        customer = str(_cell(row, mapping, 'customer', '') or '').strip().lstrip("'").strip()
        if not customer:
            skipped += 1
            continue
        entity_id = cust_by_name.get(_norm_header(customer))
        if not entity_id:
            unmatched_customers.add(customer)
            errors.append({'row': line_no, 'message': f'Müşteri eşleşmedi: {customer}'})
            continue
        code = str(_cell(row, mapping, 'code', '') or '').strip()
        barcode = str(_cell(row, mapping, 'barcode', '') or '').strip()
        pname = str(_cell(row, mapping, 'product', '') or '').strip()
        rec = None
        if code:
            rec = prod_by_code.get(_norm_header(code))
        if not rec and barcode:
            rec = prod_by_barcode.get(_norm_header(barcode))
        if not rec and pname:
            rec = prod_by_name.get(_norm_header(pname))
        if not rec:
            label = pname or code or barcode or '?'
            unmatched_products.add(label)
            errors.append({'row': line_no, 'message': f'Ürün eşleşmedi: {label}'})
            continue
        try:
            qty = _num(_cell(row, mapping, 'quantity', 0))
            net = _num(_cell(row, mapping, 'net', 0))
            gross = _num(_cell(row, mapping, 'gross', 0))
            vat = int(_num(_cell(row, mapping, 'vat_rate', rec['vat'] if rec['vat'] is not None else 20)))
        except ValueError as exc:
            errors.append({'row': line_no, 'message': f'Sayı hatası: {exc}'})
            continue
        if qty <= 0:
            errors.append({'row': line_no, 'message': 'Miktar sıfır veya geçersiz.'})
            continue
        iso = _to_iso_date(_cell(row, mapping, 'date'))
        if not iso:
            errors.append({'row': line_no, 'message': 'Geçersiz veya boş tarih.'})
            continue
        vat = min(max(vat, 0), 100)
        # BizimHesap raporu satır bazlıdır: aynı ürün, aynı müşteri-gün belgesinde
        # birden fazla satırda geçebilir. Belge modeli bir üründen tek satır kabul
        # ettiği için (TransactionCreate doğrulaması) bu satırlar BİRLEŞTİRİLİR:
        # miktarlar ve tutarlar toplanır, birim fiyat en sonda toplam tutar/toplam
        # miktar olarak türetilir. Satır başına yuvarlayıp toplamak yerine önce
        # toplayıp sonra bölmek belge toplamını rapordaki toplamla birebir tutar.
        group = groups.setdefault((entity_id, iso), {'entity_id': entity_id, 'date': iso, 'lines': {}})
        line = group['lines'].setdefault(
            rec['id'], {'qty': ZERO, 'net': ZERO, 'gross': ZERO, 'vat': vat, 'rows': 0}
        )
        if line['rows'] and line['vat'] != vat and qty > line['qty']:
            # Aynı ürün farklı KDV oranlarıyla gelmişse ağırlıklı çoğunluk kazanır.
            line['vat'] = vat
        line['qty'] += qty
        line['net'] += net
        line['gross'] += gross
        line['rows'] += 1

    from ..schemas import TransactionCreate, TransactionItem
    from .transactions import _save

    created = lines_imported = merged_rows = 0
    try:
        for group in groups.values():
            if not group['lines']:
                continue
            items = []
            for product_id, line in group['lines'].items():
                merged_rows += line['rows'] - 1
                items.append({
                    'product_id': product_id,
                    'quantity': line['qty'],
                    'unit_price': _sales_unit_price(
                        net=line['net'], gross=line['gross'], vat_rate=line['vat'], qty=line['qty']
                    ),
                    'vat_rate': line['vat'],
                })
            payload = TransactionCreate(
                entity_id=group['entity_id'],
                transaction_date=group['date'],
                status='draft',
                payment_method='credit',
                note=SALES_IMPORT_NOTE,
                items=[TransactionItem(**item) for item in items],
            )
            _save('sale', payload, db, cid, override, commit=False)
            created += 1
            lines_imported += len(items)
        db.commit()
    except HTTPException:
        db.rollback()
        raise
    except Exception as exc:
        db.rollback()
        logger.exception('Satış içe aktarma beklenmeyen hata', extra={'path': '/api/imports/sales'})
        raise HTTPException(400, IMPORT_FAILED_MESSAGE) from exc

    return {
        'documents_created': created,
        'lines_imported': lines_imported,
        'merged_duplicate_rows': merged_rows,
        'skipped_empty': skipped,
        'unmatched_customer_count': len(unmatched_customers),
        'unmatched_customers': sorted(unmatched_customers)[:100],
        'unmatched_product_count': len(unmatched_products),
        'unmatched_products': sorted(unmatched_products)[:100],
        'errors': errors[:500],
        'total_rows': len(rows),
    }


# --- Tahsilat içe aktarma ----------------------------------------------------
# Eski sistemden geçişte cari bakiyeleri doğrulamanın tek yolu buydu: satış
# geçmişi taşınabiliyor ama tahsilatlar taşınamıyordu (BizimHesap onları tek tek
# tutmuyor), dolayısıyla her müşteri gerçekte olduğundan borçlu görünüyordu.
# Elle 50+ tahsilat girmek hem uzun hem kuruş hatasına açık.
#
# Satır başına MEVCUT tahsilat ucu (`routers.finance.add`) çağrılır; kendi SQL'i
# yazılmaz. Böylece tahsis motoru, hesap doğrulaması, denetim kaydı ve
# idempotency aynen korunur — içe aktarma ayrıcalıklı bir yol değildir.
PAYMENT_METHODS = {
    'nakit': 'cash', 'cash': 'cash',
    'kart': 'card', 'pos': 'card', 'card': 'card',
    'havale': 'bank_transfer', 'eft': 'bank_transfer', 'banka': 'bank_transfer',
    'bank_transfer': 'bank_transfer',
    'cek': 'check', 'çek': 'check', 'check': 'check',
    'senet': 'promissory_note', 'promissory_note': 'promissory_note',
}


@router.post('/payments/excel')
async def import_payments(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Cari + tarih + tutar içeren bir Excel'i tahsilat kaydına dönüştürür."""
    from .finance import add as create_payment

    if not (file.filename or '').lower().endswith(('.xlsx', '.xlsm')):
        raise HTTPException(400, 'Yalnızca .xlsx dosyası yükleyin.')
    headers, rows = _read_rows(await _read_upload_limited(file))
    mapping = _map(headers, {
        'entity': ['Cari', 'Müşteri', 'Müşteri Adı', 'İsim/Ünvan'],
        'date': ['Tarih', 'Tahsilat Tarihi'],
        'amount': ['Tutar', 'Tahsilat', 'Amount'],
        'method': ['Yöntem', 'Ödeme Yöntemi', 'Şekli'],
        'note': ['Not', 'Açıklama'],
    })
    for key, label in (('entity', 'Cari'), ('amount', 'Tutar')):
        if key not in mapping:
            raise HTTPException(400, f'{label} sütunu bulunamadı.')

    cid = company_id(request)
    by_name: dict[str, int] = {}
    for row in db.execute(
        text('SELECT id,name FROM customers WHERE company_id=:cid'), {'cid': cid}
    ).mappings():
        by_name.setdefault(_norm_header(row['name']), int(row['id']))

    created = 0
    errors: list[dict[str, Any]] = []
    unmatched: set[str] = set()
    for line_no, row in enumerate(rows, start=2):
        name = str(_cell(row, mapping, 'entity', '') or '').strip()
        if not name:
            continue
        entity_id = by_name.get(_norm_header(name))
        if not entity_id:
            unmatched.add(name)
            errors.append({'row': line_no, 'message': f'Cari eşleşmedi: {name}'})
            continue
        try:
            amount = decimal_value(_cell(row, mapping, 'amount', 0))
        except (ValueError, ArithmeticError):
            errors.append({'row': line_no, 'message': 'Tutar sayı değil.'})
            continue
        if amount <= ZERO:
            # Tahsilat negatif olamaz (PaymentCreate: amount gt=0). Ters yön
            # açılış bakiyesiyle çözülür; burada sessizce ters çevirmek bakiyeyi
            # yanlış yöne taşır.
            errors.append({'row': line_no, 'message': 'Tutar sıfır veya negatif.'})
            continue
        iso = _to_iso_date(_cell(row, mapping, 'date'))
        if not iso:
            errors.append({'row': line_no, 'message': 'Tarih okunamadı.'})
            continue
        raw_method = str(_cell(row, mapping, 'method', '') or '').strip().lower()
        method = PAYMENT_METHODS.get(raw_method, 'cash')
        try:
            create_payment(
                PaymentCreate(
                    entity_type='customer',
                    entity_id=entity_id,
                    amount=amount,
                    payment_date=iso,
                    note=str(_cell(row, mapping, 'note', '') or '').strip() or None,
                    payment_method=method,
                ),
                request,
                db,
            )
            created += 1
        except HTTPException as exc:
            errors.append({'row': line_no, 'message': str(exc.detail)})
        except Exception as exc:  # noqa: BLE001 - satır hatası tüm dosyayı düşürmesin
            db.rollback()
            logger.exception('Tahsilat içe aktarma satır hatası')
            errors.append({'row': line_no, 'message': f'Kaydedilemedi: {exc}'})
    return {
        'inserted': created,
        'updated': 0,
        'unmatched_entities': sorted(unmatched)[:100],
        'errors': errors[:500],
        'total_rows': len(rows),
    }
