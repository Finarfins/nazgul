from __future__ import annotations

import hashlib
import io
import json
import logging
import re
import zipfile
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any, Literal

from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, Field, model_validator
from sqlalchemy import bindparam, text
from sqlalchemy.exc import IntegrityError, OperationalError
from sqlalchemy.orm import Session

from ..activity_log import log_activity
from ..auth import has_permission
from ..change_history import record_change
from ..db import get_db
from ..late_fee_schemas import RateOut
from ..pos_contracts import MoneyOut
from ..supplier_price_bridge import (
    canonical_price_digest,
    classify_deviation,
    collect_reason_codes,
    deviation_percent,
    normalize_price_cell,
    split_supplier_code,
    vat_details,
)
from ..supplier_price_pdf import PdfSectionOverlapError, extract_pdf_rows
from ..tenancy import company_id

router = APIRouter(prefix="/supplier-prices", tags=["supplier-price-bridge"])
logger = logging.getLogger("yerel_hesap.supplier_price_bridge")

MAX_UPLOAD = 25 * 1024 * 1024
MAX_ENTRIES = 512
MAX_ENTRY_BYTES = 100 * 1024 * 1024
MAX_TOTAL_BYTES = 300 * 1024 * 1024
MAX_ROWS = 50_000
MAX_SHEETS = 200
READ_CHUNK = 64 * 1024


class PageSectionInput(BaseModel):
    page_start: int | None = Field(default=None, ge=1)
    page_end: int | None = Field(default=None, ge=1)
    header_signature: str | None = Field(default=None, min_length=1, max_length=255)
    column_map: dict[str, int]

    @model_validator(mode="after")
    def validate_selector(self) -> "PageSectionInput":
        has_range = self.page_start is not None or self.page_end is not None
        if has_range and (
            self.page_start is None
            or self.page_end is None
            or self.page_end < self.page_start
        ):
            raise ValueError("Sayfa aralığı başlangıç ve bitiş ile geçerli olmalıdır")
        if not has_range and self.header_signature is None:
            raise ValueError("Sayfa aralığı veya başlık imzası zorunludur")
        required = {"code", "price_cash", "price_term"}
        if not required.issubset(self.column_map):
            raise ValueError("Kod, peşin ve vadeli fiyat kolonları zorunludur")
        return self


class ProfileInput(BaseModel):
    supplier_id: int = Field(gt=0)
    name: str = Field(min_length=1, max_length=160)
    source_type: Literal["xlsx", "pdf"] = "xlsx"
    sheet_selector: str = Field(default=".*", min_length=1, max_length=255)
    header_row_strategy: str = Field(default="detect", pattern=r"^(fixed:\d+|detect)$")
    column_map: dict[str, str] = Field(default_factory=dict)
    page_sections: list[PageSectionInput] = Field(default_factory=list, max_length=100)
    currency_mode: str = Field(pattern=r"^(inline|column|fixed:(TRY|EUR|USD))$")
    term_days: int = Field(ge=0, le=3650)
    vat_source: str = Field(pattern=r"^(indeks_sheet|column|fixed:\d+(\.\d+)?)$")
    warning_threshold: Decimal = Field(default=Decimal("25"), ge=0)
    blocked_threshold: Decimal = Field(default=Decimal("100"), ge=0)

    @model_validator(mode="after")
    def validate_source_mapping(self) -> "ProfileInput":
        required = {"code", "price_cash", "price_term"}
        if self.source_type == "xlsx" and not required.issubset(self.column_map):
            raise ValueError("Kod, peşin ve vadeli fiyat kolonları zorunludur")
        if self.source_type == "pdf" and not self.page_sections:
            raise ValueError("PDF profili en az bir sayfa bölümü içermelidir")
        for index, section in enumerate(self.page_sections):
            for other_index, other in enumerate(
                self.page_sections[index + 1 :], start=index + 1
            ):
                ranges_overlap = (
                    section.page_start is not None
                    and other.page_start is not None
                    and section.page_start <= other.page_end
                    and other.page_start <= section.page_end
                )
                signatures_overlap = (
                    section.header_signature is not None
                    and other.header_signature is not None
                    and section.header_signature.casefold()
                    == other.header_signature.casefold()
                )
                if ranges_overlap or signatures_overlap:
                    raise ValueError(
                        f"PDF bölümleri {index + 1} ve {other_index + 1} çakışıyor"
                    )
        return self


class ProfileSummaryOutput(BaseModel):
    id: int
    supplier_id: int
    supplier_name: str
    name: str
    source_type: Literal["xlsx", "pdf"]
    currency_mode: str
    term_days: int
    vat_source: str
    warning_threshold: str
    blocked_threshold: str
    created_at: datetime
    import_count: int


class ProfileDetailOutput(ProfileSummaryOutput):
    sheet_selector: str
    header_row_strategy: str
    column_map: dict[str, str]
    page_sections: list[PageSectionInput]
    updated_at: datetime
    created_by: int | None


class ApplyInput(BaseModel):
    expected_revision: str
    report_digest: str


class OverrideInput(BaseModel):
    reason: str = Field(min_length=3, max_length=1000)


class ImportReportLine(BaseModel):
    line_no: int
    supplier_code: str | None
    description: str | None
    unit: str | None
    price_cash: RateOut | None
    price_term: RateOut | None
    currency: str | None
    vat_rate: Decimal | None
    part_id: int | None
    match_source: str | None
    old_price_cash: RateOut | None
    old_price_term: RateOut | None
    old_currency: str | None
    deviation_pct: MoneyOut | None
    state: str
    raw: dict[str, Any]
    override_by: int | None
    override_at: datetime | None
    override_reason: str | None
    status_note: str | None
    reason_code: str | None
    reason_codes: dict[str, str | None]
    vat_included: bool
    has_tier: bool
    change_kind: str


class ImportReport(BaseModel):
    id: int
    supplier_id: int
    profile_id: int
    status: str
    source_filename: str
    parsed_row_count: int
    matched_count: int
    unmatched_count: int
    warning_count: int
    blocked_count: int
    revision: int | None
    report_digest: str
    after_digest: str | None
    applied_at: datetime | None
    reverted_at: datetime | None
    created_at: datetime
    expected_revision: str
    review_count: int
    state_summary: dict[str, int]
    state_total: int
    extraction_report: Any | None
    change_summary: dict[str, int]
    tier_count: int
    lines: list[ImportReportLine]


class XrefInput(BaseModel):
    supplier_id: int = Field(gt=0)
    supplier_code: str = Field(min_length=1, max_length=255)
    part_id: int = Field(gt=0)


def _now() -> datetime:
    return datetime.now(timezone.utc)


def _actor(request: Request) -> dict[str, Any]:
    return getattr(request.state, "user", {}) or {}


def _require(request: Request, permission: str) -> None:
    if not has_permission(str(_actor(request).get("role", "")), permission):
        raise HTTPException(403, "Bu işlem için yetkiniz yok")


async def _read_upload(file: UploadFile) -> bytes:
    if file.size is not None and file.size > MAX_UPLOAD:
        raise HTTPException(400, "Dosya boyutu 25 MB sınırını aşıyor")
    content = bytearray()
    while chunk := await file.read(READ_CHUNK):
        content.extend(chunk)
        if len(content) > MAX_UPLOAD:
            raise HTTPException(400, "Dosya boyutu 25 MB sınırını aşıyor")
    return bytes(content)


def _validate_upload_type(file: UploadFile, content: bytes, source_type: str) -> None:
    suffix = Path(file.filename or "").suffix.lower()
    if source_type == "pdf":
        if suffix != ".pdf" or not content.startswith(b"%PDF-"):
            raise HTTPException(
                400,
                "Dosya türü profil ile uyuşmuyor: PDF profili geçerli bir .pdf dosyası gerektirir.",
            )
        return
    if suffix not in {".xlsx", ".xlsm"} or not content.startswith(b"PK\x03\x04"):
        raise HTTPException(
            400,
            "Dosya türü profil ile uyuşmuyor: Excel profili geçerli bir .xlsx veya .xlsm dosyası gerektirir.",
        )


def _validate_archive(content: bytes) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            entries = archive.infolist()
            if len(entries) > MAX_ENTRIES:
                raise HTTPException(400, "Excel ZIP entry sınırını aşıyor")
            total = 0
            buffer = bytearray(READ_CHUNK)
            for entry in entries:
                opened = 0
                with archive.open(entry) as stream:
                    while count := stream.readinto(buffer):
                        opened += count
                        total += count
                        if opened > MAX_ENTRY_BYTES or total > MAX_TOTAL_BYTES:
                            raise HTTPException(400, "Excel açılmış veri sınırını aşıyor")
    except zipfile.BadZipFile as exc:
        raise HTTPException(400, "Geçersiz Excel dosyası") from exc


def _selected_sheets(workbook: Any, selector: str) -> list[str]:
    if len(workbook.sheetnames) > MAX_SHEETS:
        raise HTTPException(400, "Excel en fazla 200 sheet içerebilir")
    if selector in workbook.sheetnames:
        return [selector]
    try:
        pattern = re.compile(selector)
    except re.error as exc:
        raise HTTPException(400, "Profil sheet regex'i geçersiz") from exc
    selected = [name for name in workbook.sheetnames if pattern.fullmatch(name)]
    if not selected:
        raise HTTPException(400, "Profil hiçbir sheet ile eşleşmedi")
    return selected


def _header_row(rows: Any, strategy: str, required_headers: set[str]) -> tuple[int, list[str]]:
    if strategy.startswith("fixed:"):
        number = int(strategy.split(":", 1)[1])
        values = next(rows.iter_rows(min_row=number, max_row=number, values_only=True))
        return number, [str(value or "").strip() for value in values]
    for number, values in enumerate(rows.iter_rows(values_only=True), start=1):
        headers = [str(value or "").strip() for value in values]
        if required_headers.issubset(set(headers)):
            return number, headers
        if number >= 100:
            break
    raise HTTPException(400, "Başlık satırı bulunamadı")


def _json_value(value: Any) -> Any:
    return value if value is None or isinstance(value, (str, int, bool)) else str(value)


def _vat(profile: dict[str, Any], row: dict[str, Any]) -> tuple[Decimal, bool]:
    source = str(profile["vat_source"])
    if source.startswith("fixed:"):
        return Decimal(source.split(":", 1)[1]).quantize(Decimal("0.01")), False
    if source == "column":
        column = profile["column_map"].get("vat_rate")
        return vat_details(row.get(column))
    return Decimal("0.00"), False


def _currency(profile: dict[str, Any], row: dict[str, Any], inline: str | None) -> str:
    mode = str(profile["currency_mode"])
    if mode.startswith("fixed:"):
        return mode.split(":", 1)[1]
    if mode == "column":
        value = str(row.get(profile["column_map"].get("currency"), "")).upper().strip()
        return {"TL": "TRY"}.get(value, value)
    return inline or ""


def _resolve_part(db: Session, cid: int, supplier_id: int, code: str) -> tuple[int | None, str | None]:
    xref = db.execute(
        text("""SELECT part_id FROM supplier_part_xrefs
                WHERE company_id=:cid AND supplier_id=:sid AND supplier_code=:code"""),
        {"cid": cid, "sid": supplier_id, "code": code},
    ).scalar_one_or_none()
    if xref is not None:
        return int(xref), "xref"
    rows = db.execute(
        text("""SELECT id FROM products WHERE company_id=:cid
                AND (product_code=:code OR oem_number=:code OR alternative_oem=:code)
                AND COALESCE(active,TRUE)=TRUE ORDER BY id"""),
        {"cid": cid, "code": code},
    ).scalars().all()
    return (int(rows[0]), "exact_oem") if len(rows) == 1 else (None, None)


class ExtractionSuspectError(ValueError):
    pass


class MissingPriceError(ValueError):
    pass


def _parse_xlsx(
    db: Session, cid: int, import_id: int, content: bytes, profile: dict[str, Any]
) -> None:
    _validate_archive(content)
    workbook = load_workbook(
        io.BytesIO(content), read_only=True, data_only=True, keep_links=False
    )
    mapping = dict(profile["column_map"])
    required = {mapping["code"], mapping["price_cash"], mapping["price_term"]}
    line_no = parsed = matched = unmatched = warnings = blocked = 0
    digest_rows: list[str] = []
    for sheet_name in _selected_sheets(workbook, str(profile["sheet_selector"])):
        sheet = workbook[sheet_name]
        header_no, headers = _header_row(sheet, str(profile["header_row_strategy"]), required)
        if sheet.max_row is not None and sheet.max_row - header_no > MAX_ROWS:
            raise HTTPException(400, "Excel 50.000 satır sınırını aşıyor")
        for values in sheet.iter_rows(min_row=header_no + 1, values_only=True):
            if not any(value not in (None, "") for value in values):
                continue
            parsed += 1
            if parsed > MAX_ROWS:
                raise HTTPException(400, "Excel 50.000 satır sınırını aşıyor")
            line_no += 1
            raw = {headers[index]: _json_value(value) for index, value in enumerate(values) if index < len(headers)}
            code, description = split_supplier_code(raw.get(mapping["code"]), raw.get(mapping.get("description")))
            state = "SKIPPED"
            cash = term = vat = deviation = None
            vat_included = False
            currency = ""
            part_id = match_source = None
            old = None
            status_note = str(raw.get(mapping.get("status_note")) or "").strip() or None
            reason_code = None
            reason_codes: dict[str, str | None] = {}
            try:
                if raw.pop("_extraction_suspect", False):
                    raise ExtractionSuspectError
                if not code:
                    raise ValueError("Kod boş")
                if all(
                    raw.get(mapping[key]) is None
                    or str(raw.get(mapping[key])).strip() == ""
                    for key in ("price_cash", "price_term")
                ):
                    state = "REVIEW"
                    reason_code = "MISSING_PRICE"
                    reason_codes["row"] = reason_code
                    raise MissingPriceError
                cash_cell = normalize_price_cell(raw.get(mapping["price_cash"]))
                term_cell = normalize_price_cell(raw.get(mapping["price_term"]))
                note_cell = normalize_price_cell(status_note) if status_note else None
                reason_codes, reason_code = collect_reason_codes(
                    cash_cell, term_cell, note_cell
                )
                price_notes = [
                    cell.status_note
                    for cell in (cash_cell, term_cell)
                    if cell.status_note
                ]
                status_note = "\n".join(dict.fromkeys([*price_notes, status_note] if status_note else price_notes)) or None
                if reason_code:
                    state = "REVIEW"
                    continue_parse = False
                else:
                    continue_parse = True
                cash, term = cash_cell.amount, term_cell.amount
                cash_currency, term_currency = cash_cell.currency, term_cell.currency
                currency = _currency(profile, raw, cash_currency or term_currency)
                if continue_parse and (currency not in {"TRY", "EUR", "USD"} or (
                    cash_currency and term_currency and cash_currency != term_currency
                )):
                    reason_code = "CURRENCY_MISMATCH"
                    reason_codes["row"] = reason_code
                    state = "REVIEW"
                    continue_parse = False
                vat, vat_included = _vat(profile, raw)
                if continue_parse:
                    part_id, match_source = _resolve_part(db, cid, int(profile["supplier_id"]), code)
                if continue_parse and part_id is None:
                    state = "UNMATCHED"
                    unmatched += 1
                elif continue_parse:
                    old = db.execute(
                        text("""SELECT price_cash,price_term,currency,vat_rate
                                FROM supplier_part_prices
                                WHERE company_id=:cid AND supplier_id=:sid AND part_id=:pid"""),
                        {"cid": cid, "sid": profile["supplier_id"], "pid": part_id},
                    ).mappings().first()
                    changes = [
                        deviation_percent(old.get(key) if old else None, value)
                        for key, value in (("price_cash", cash), ("price_term", term))
                    ]
                    deviation = max(changes)
                    state = classify_deviation(
                        deviation,
                        Decimal(str(profile["warning_threshold"])),
                        Decimal(str(profile["blocked_threshold"])),
                    )
                    matched += 1
                    warnings += state == "WARNING"
                    blocked += state == "BLOCKED"
            except ExtractionSuspectError:
                state = "REVIEW"
                reason_code = "EXTRACTION_SUSPECT"
                reason_codes["row"] = reason_code
            except MissingPriceError:
                pass
            except (ValueError, KeyError, ArithmeticError) as exc:
                state = "SKIPPED"
                reason_code = {
                    "Kod boş": "STRUCTURAL_ROW",
                }.get(str(exc), "INVALID_ROW")
                reason_codes["row"] = reason_code
            raw["_normalization"] = {
                "status_note": status_note,
                "reason_code": reason_code,
                "reason_codes": reason_codes,
                "vat_included": vat_included,
            }
            params = {
                "cid": cid, "iid": import_id, "line": line_no,
                "raw": json.dumps(raw, ensure_ascii=False), "code": code or None,
                "description": description, "unit": raw.get(mapping.get("unit")),
                "cash": cash, "term": term, "currency": currency or None, "vat": vat,
                "part_id": part_id, "match_source": match_source,
                "old_cash": old.get("price_cash") if old else None,
                "old_term": old.get("price_term") if old else None,
                "old_currency": old.get("currency") if old else None,
                "old_vat": old.get("vat_rate") if old else None,
                "deviation": deviation, "state": state,
            }
            db.execute(text("""INSERT INTO supplier_price_import_lines(
                company_id,import_id,line_no,raw,supplier_code,description,unit,
                price_cash,price_term,currency,vat_rate,part_id,match_source,
                old_price_cash,old_price_term,old_currency,old_vat_rate,deviation_pct,state)
                VALUES(:cid,:iid,:line,:raw,:code,:description,:unit,:cash,:term,:currency,
                :vat,:part_id,:match_source,:old_cash,:old_term,:old_currency,:old_vat,
                :deviation,:state)"""), params)
            digest_rows.append("|".join(str(params[key]) for key in ("line", "code", "cash", "term", "currency", "state", "part_id")))
    report_digest = hashlib.sha256("\n".join(digest_rows).encode()).hexdigest()
    db.execute(text("""UPDATE supplier_price_imports SET status='DRY_RUN_READY',
        parsed_row_count=:parsed,matched_count=:matched,unmatched_count=:unmatched,
        warning_count=:warnings,blocked_count=:blocked,report_digest=:digest
        WHERE company_id=:cid AND id=:iid"""),
        {"parsed": parsed, "matched": matched, "unmatched": unmatched, "warnings": warnings,
         "blocked": blocked, "digest": report_digest, "cid": cid, "iid": import_id})


def _parse_pdf(
    db: Session, cid: int, import_id: int, content: bytes, profile: dict[str, Any]
) -> None:
    sections = list(profile.get("page_sections") or [])
    headers = sorted(
        {
            field
            for section in sections
            for field in section["column_map"]
        }
        | {"_page_number", "_pdf_row_number", "_extraction_suspect"}
    )
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("PDF")
    sheet.append(headers)
    extraction = extract_pdf_rows(content, sections)
    for row in extraction.rows:
        values = {
            **row.values,
            "_page_number": row.page_number,
            "_pdf_row_number": row.row_number,
            "_extraction_suspect": row.extraction_suspect,
        }
        sheet.append([values.get(header) for header in headers])
    output = io.BytesIO()
    workbook.save(output)
    pdf_profile = {
        **profile,
        "sheet_selector": "PDF",
        "header_row_strategy": "fixed:1",
        "column_map": {
            field: field
            for field in headers
            if not field.startswith("_")
        },
    }
    _parse_xlsx(db, cid, import_id, output.getvalue(), pdf_profile)
    db.execute(
        text(
            """UPDATE supplier_price_imports SET error_detail=:report
            WHERE company_id=:cid AND id=:iid"""
        ),
        {
            "report": json.dumps({"extraction": extraction.report}),
            "cid": cid,
            "iid": import_id,
        },
    )


# Kept for focused F1 unit tests and callers that exercise the XLSX parser directly.
_parse = _parse_xlsx


_PROFILE_IN_USE = "Bu profille yapılmış içe aktarımlar var; profil silinemez."
_PROFILE_LOCKED = "Kayıt geçici olarak kilitli, lütfen tekrar deneyin."
# Seconds. Short on purpose: the contended writer is a single import INSERT,
# not a long batch, so the caller should come straight back.
_PROFILE_LOCK_RETRY_AFTER = "1"


def _profile_not_found() -> HTTPException:
    return HTTPException(404, "Eşleme profili bulunamadı")


def _profile_import_count(db: Session, cid: int, profile_id: int) -> int:
    """Tenant-scoped count of imports referencing ``profile_id``."""
    return db.execute(
        text(
            """SELECT COUNT(*) FROM supplier_price_imports
               WHERE company_id=:cid AND profile_id=:id"""
        ),
        {"cid": cid, "id": profile_id},
    ).scalar_one()


def _profile_json(value: Any, default: Any) -> Any:
    if value is None:
        return default
    return json.loads(value) if isinstance(value, str) else value


def _profile_output(row: Any, *, include_mapping: bool = False) -> dict[str, Any]:
    result = dict(row)
    result["warning_threshold"] = f"{Decimal(str(result['warning_threshold'])):.2f}"
    result["blocked_threshold"] = f"{Decimal(str(result['blocked_threshold'])):.2f}"
    if include_mapping:
        result["column_map"] = _profile_json(result.get("column_map"), {})
        result["page_sections"] = _profile_json(result.get("page_sections"), [])
    return result


@router.get("/profiles", response_model=list[ProfileSummaryOutput])
def list_profiles(request: Request, db: Session = Depends(get_db)):
    _require(request, "supplier_prices.view")
    cid = company_id(request)
    rows = db.execute(
        text(
            """SELECT p.id,p.supplier_id,s.name AS supplier_name,p.name,p.source_type,
                      p.currency_mode,p.term_days,p.vat_source,p.warning_threshold,
                      p.blocked_threshold,p.created_at,
                      (SELECT COUNT(*) FROM supplier_price_imports i
                       WHERE i.company_id=p.company_id AND i.profile_id=p.id) AS import_count
               FROM supplier_import_profiles p
               JOIN suppliers s ON s.company_id=p.company_id AND s.id=p.supplier_id
               WHERE p.company_id=:cid
               ORDER BY p.name ASC,p.id ASC"""
        ),
        {"cid": cid},
    ).mappings().all()
    return [_profile_output(row) for row in rows]


@router.get("/profiles/{profile_id}", response_model=ProfileDetailOutput)
def get_profile(
    profile_id: int, request: Request, db: Session = Depends(get_db)
):
    _require(request, "supplier_prices.view")
    row = db.execute(
        text(
            """SELECT p.id,p.supplier_id,s.name AS supplier_name,p.name,p.source_type,
                      p.sheet_selector,p.header_row_strategy,p.column_map,p.page_sections,
                      p.currency_mode,p.term_days,p.vat_source,p.warning_threshold,
                      p.blocked_threshold,p.created_at,p.updated_at,p.created_by,
                      (SELECT COUNT(*) FROM supplier_price_imports i
                       WHERE i.company_id=p.company_id AND i.profile_id=p.id) AS import_count
               FROM supplier_import_profiles p
               JOIN suppliers s ON s.company_id=p.company_id AND s.id=p.supplier_id
               WHERE p.company_id=:cid AND p.id=:id
               """
        ),
        {"cid": company_id(request), "id": profile_id},
    ).mappings().first()
    if not row:
        raise _profile_not_found()
    return _profile_output(row, include_mapping=True)


@router.post("/profiles", status_code=201)
def create_profile(payload: ProfileInput, request: Request, db: Session = Depends(get_db)):
    _require(request, "supplier_prices.import")
    cid, actor = company_id(request), _actor(request)
    if payload.blocked_threshold < payload.warning_threshold:
        raise HTTPException(400, "Blok eşiği uyarı eşiğinden küçük olamaz")
    supplier = db.execute(text("SELECT id FROM suppliers WHERE company_id=:cid AND id=:id"),
                          {"cid": cid, "id": payload.supplier_id}).first()
    if not supplier:
        raise HTTPException(404, "Tedarikçi bulunamadı")
    values = payload.model_dump()
    values.update(
        cid=cid,
        uid=actor.get("id"),
        now=_now(),
        column_map=json.dumps(payload.column_map),
        page_sections=json.dumps(
            [section.model_dump() for section in payload.page_sections]
        ),
    )
    profile_id = db.execute(text("""INSERT INTO supplier_import_profiles(
        company_id,supplier_id,name,source_type,sheet_selector,header_row_strategy,column_map,page_sections,
        currency_mode,term_days,vat_source,warning_threshold,blocked_threshold,
        created_at,updated_at,created_by)
        VALUES(:cid,:supplier_id,:name,:source_type,:sheet_selector,:header_row_strategy,:column_map,:page_sections,
        :currency_mode,:term_days,:vat_source,:warning_threshold,:blocked_threshold,
        :now,:now,:uid) RETURNING id"""), values).scalar_one()
    if (
        payload.warning_threshold != Decimal("25")
        or payload.blocked_threshold != Decimal("100")
    ):
        record_change(
            db,
            request,
            company_id=cid,
            entity_type="supplier_import_profile",
            entity_id=int(profile_id),
            action="threshold_override",
            before={"warning_threshold": "25", "blocked_threshold": "100"},
            after={
                "warning_threshold": str(payload.warning_threshold),
                "blocked_threshold": str(payload.blocked_threshold),
            },
        )
    db.commit()
    return {"id": profile_id}


@router.put("/profiles/{profile_id}", response_model=ProfileDetailOutput)
def update_profile(
    profile_id: int,
    payload: ProfileInput,
    request: Request,
    db: Session = Depends(get_db),
):
    _require(request, "supplier_prices.import")
    cid, actor = company_id(request), _actor(request)
    if payload.blocked_threshold < payload.warning_threshold:
        raise HTTPException(400, "Blok eşiği uyarı eşiğinden küçük olamaz")
    current = db.execute(
        text(
            """SELECT id,warning_threshold,blocked_threshold
               FROM supplier_import_profiles WHERE company_id=:cid AND id=:id"""
        ),
        {"cid": cid, "id": profile_id},
    ).mappings().first()
    if not current:
        raise _profile_not_found()
    supplier = db.execute(
        text("SELECT id FROM suppliers WHERE company_id=:cid AND id=:id"),
        {"cid": cid, "id": payload.supplier_id},
    ).first()
    if not supplier:
        raise HTTPException(404, "Tedarikçi bulunamadı")
    values = payload.model_dump()
    values.update(
        cid=cid,
        id=profile_id,
        uid=actor.get("id"),
        now=_now(),
        column_map=json.dumps(payload.column_map),
        page_sections=json.dumps(
            [section.model_dump() for section in payload.page_sections]
        ),
    )
    try:
        db.execute(
            text(
                """UPDATE supplier_import_profiles SET
                   supplier_id=:supplier_id,name=:name,source_type=:source_type,
                   sheet_selector=:sheet_selector,header_row_strategy=:header_row_strategy,
                   column_map=:column_map,page_sections=:page_sections,
                   currency_mode=:currency_mode,term_days=:term_days,vat_source=:vat_source,
                   warning_threshold=:warning_threshold,blocked_threshold=:blocked_threshold,
                   updated_at=:now
                   WHERE company_id=:cid AND id=:id"""
            ),
            values,
        )
        if (
            Decimal(str(current["warning_threshold"])) != payload.warning_threshold
            or Decimal(str(current["blocked_threshold"])) != payload.blocked_threshold
        ):
            record_change(
                db,
                request,
                company_id=cid,
                entity_type="supplier_import_profile",
                entity_id=profile_id,
                action="threshold_override",
                before={
                    "warning_threshold": str(current["warning_threshold"]),
                    "blocked_threshold": str(current["blocked_threshold"]),
                },
                after={
                    "warning_threshold": str(payload.warning_threshold),
                    "blocked_threshold": str(payload.blocked_threshold),
                },
            )
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(
            409, "Bu tedarikçi için aynı adlı bir eşleme profili zaten var"
        ) from exc
    return get_profile(profile_id, request, db)


@router.delete("/profiles/{profile_id}")
def delete_profile(
    profile_id: int, request: Request, db: Session = Depends(get_db)
):
    _require(request, "supplier_prices.import")
    cid = company_id(request)
    # Lock the profile row before counting its imports. ``supplier_price_imports``
    # carries a FK to this row, so on PostgreSQL a concurrent import INSERT takes
    # a FOR KEY SHARE lock on it -- which conflicts with FOR UPDATE. The racing
    # writer is therefore serialised against us: either it commits first and the
    # COUNT below sees its row, or it waits until our DELETE has committed and
    # then fails its own FK check. SQLite has no row locks, so the suffix is
    # empty there and the IntegrityError net below carries the guarantee.
    lock_clause = " FOR UPDATE" if db.get_bind().dialect.name == "postgresql" else ""
    exists = db.execute(
        text(
            "SELECT id FROM supplier_import_profiles "
            f"WHERE company_id=:cid AND id=:id{lock_clause}"
        ),
        {"cid": cid, "id": profile_id},
    ).first()
    if not exists:
        raise _profile_not_found()
    import_count = _profile_import_count(db, cid, profile_id)
    if import_count:
        # Release the row lock before answering, so a blocked importer can
        # proceed instead of waiting for request teardown.
        db.rollback()
        raise HTTPException(409, _PROFILE_IN_USE)
    try:
        db.execute(
            text(
                "DELETE FROM supplier_import_profiles WHERE company_id=:cid AND id=:id"
            ),
            {"cid": cid, "id": profile_id},
        )
        db.commit()
    except IntegrityError as exc:
        # Safety net kept even when the lock was taken: it covers SQLite, which
        # has no row lock, and any FK violation surfaced only at COMMIT. The
        # cause is the same either way, so the caller gets the same 409.
        db.rollback()
        raise HTTPException(409, _PROFILE_IN_USE) from exc
    except OperationalError as exc:
        # SQLite has no row lock, so the check-then-delete window stays open and
        # a concurrent writer can hold the database write lock while we try to
        # upgrade. The DELETE then fails busy (SQLITE_BUSY / BUSY_SNAPSHOT),
        # which SQLAlchemy raises as OperationalError -- an uncontrolled 500
        # before this branch existed.
        #
        # "Busy" is NOT "in use": answering 409 here would invent a business
        # rule violation that may not exist. So measure the real cause on a
        # clean transaction before answering.
        db.rollback()
        try:
            import_count = _profile_import_count(db, cid, profile_id)
        except OperationalError:
            # Still locked out. We cannot prove the profile is in use, so we
            # must not claim that it is.
            import_count = 0
        db.rollback()
        if import_count:
            raise HTTPException(409, _PROFILE_IN_USE) from exc
        # Previously this collision reached the middleware and was logged there
        # as an unhandled 500. Answering it properly must not also make it
        # invisible: a database that keeps returning busy is an operational
        # signal, not a routine client error.
        logger.warning(
            "Eşleme profili silinemedi: veritabanı meşgul",
            extra={"company_id": cid, "profile_id": profile_id},
        )
        raise HTTPException(
            503, _PROFILE_LOCKED, headers={"Retry-After": _PROFILE_LOCK_RETRY_AFTER}
        ) from exc
    return {"ok": True}


@router.post("/imports", status_code=202)
async def create_import(request: Request, profile_id: int = Form(...), file: UploadFile = File(...),
                        db: Session = Depends(get_db)):
    _require(request, "supplier_prices.import")
    cid, actor = company_id(request), _actor(request)
    profile_row = db.execute(text("""SELECT * FROM supplier_import_profiles
        WHERE company_id=:cid AND id=:id"""), {"cid": cid, "id": profile_id}).mappings().first()
    if not profile_row:
        raise HTTPException(400, "Geçerli eşleme profili zorunludur")
    content = await _read_upload(file)
    _validate_upload_type(file, content, str(profile_row["source_type"]))
    digest = hashlib.sha256(content).hexdigest()
    already_applied = db.execute(
        text("""SELECT id FROM supplier_price_imports
                WHERE company_id=:cid AND source_sha256=:digest AND status='APPLIED'"""),
        {"cid": cid, "digest": digest},
    ).first()
    if already_applied:
        raise HTTPException(409, "Bu dosya daha önce uygulandı")
    profile = dict(profile_row)
    if isinstance(profile["column_map"], str):
        profile["column_map"] = json.loads(profile["column_map"])
    if isinstance(profile.get("page_sections"), str):
        profile["page_sections"] = json.loads(profile["page_sections"])
    try:
        import_id = db.execute(text("""INSERT INTO supplier_price_imports(
            company_id,supplier_id,profile_id,status,source_filename,source_sha256,
            source_bytes,created_at,created_by)
            VALUES(:cid,:sid,:pid,'PARSING',:name,:digest,:content,:now,:uid) RETURNING id"""),
            {"cid": cid, "sid": profile["supplier_id"], "pid": profile_id,
             "name": (file.filename or "import.xlsx")[:255], "digest": digest,
             "content": len(content), "now": _now(), "uid": actor.get("id")}).scalar_one()
        parser = _parse_pdf if profile["source_type"] == "pdf" else _parse_xlsx
        parser(db, cid, int(import_id), content, profile)
        log_activity(db, cid, actor.get("id"), "supplier_price.import_created",
                     "supplier_price_import", int(import_id), "Tedarikçi fiyat importu oluşturuldu",
                     {"supplier_id": profile["supplier_id"], "sha256": digest})
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(409, "Bu dosya için etkin bir import zaten var") from exc
    except PdfSectionOverlapError as exc:
        db.rollback()
        raise HTTPException(400, str(exc)) from exc
    except Exception as exc:
        db.rollback()
        if isinstance(exc, HTTPException):
            raise
        logger.exception("Tedarikçi fiyat dosyası ayrıştırılamadı")
        raise HTTPException(400, "Tedarikçi fiyat dosyası ayrıştırılamadı") from exc
    return {"id": import_id, "status": "DRY_RUN_READY"}


@router.get("/imports")
def list_imports(request: Request, db: Session = Depends(get_db)):
    _require(request, "supplier_prices.view")
    cid = company_id(request)
    rows = db.execute(
        text(
            """SELECT id,status,source_filename,parsed_row_count,created_at
            FROM supplier_price_imports
            WHERE company_id=:cid
            ORDER BY created_at DESC,id DESC
            LIMIT 50"""
        ),
        {"cid": cid},
    ).mappings().all()
    return [dict(row) for row in rows]


@router.get("/imports/{import_id}", response_model=ImportReport)
def get_import(import_id: int, request: Request, db: Session = Depends(get_db)):
    _require(request, "supplier_prices.view")
    cid = company_id(request)
    item = db.execute(text("""SELECT id,supplier_id,profile_id,status,source_filename,
        parsed_row_count,matched_count,unmatched_count,warning_count,blocked_count,
        revision,report_digest,after_digest,error_detail,applied_at,reverted_at,created_at
        FROM supplier_price_imports WHERE company_id=:cid AND id=:id"""),
        {"cid": cid, "id": import_id}).mappings().first()
    if not item:
        raise HTTPException(404, "Import bulunamadı")
    lines = db.execute(text("""SELECT line_no,supplier_code,description,unit,price_cash,
        price_term,currency,vat_rate,part_id,match_source,old_price_cash,old_price_term,
        old_currency,deviation_pct,state,raw,override_by,override_at,override_reason
        FROM supplier_price_import_lines WHERE company_id=:cid AND import_id=:id
        ORDER BY line_no"""), {"cid": cid, "id": import_id}).mappings().all()
    report_lines = [_report_line(row) for row in lines]
    state_summary = {
        state: sum(line["state"] == state for line in report_lines)
        for state in ("OK", "WARNING", "BLOCKED", "UNMATCHED", "REVIEW", "SKIPPED")
    }
    if sum(state_summary.values()) != item["parsed_row_count"]:
        raise HTTPException(500, "Import satır dağılımı parsed_row_count ile uyuşmuyor")
    detail = item["error_detail"]
    if isinstance(detail, str):
        detail = json.loads(detail)
    return {
        **{key: value for key, value in dict(item).items() if key != "error_detail"},
        "expected_revision": item["report_digest"],
        "review_count": state_summary["REVIEW"],
        "state_summary": state_summary,
        "state_total": sum(state_summary.values()),
        "extraction_report": (detail or {}).get("extraction"),
        "change_summary": {
            kind: sum(line["change_kind"] == kind for line in report_lines)
            for kind in (
                "NEW_PRICE",
                "CHANGED_PRICE",
                "UNCHANGED",
                "UNMATCHED",
                "SKIPPED",
                "REVIEW",
            )
        },
        "tier_count": sum(line["has_tier"] for line in report_lines),
        "lines": report_lines,
    }


def _report_line(row: Any) -> dict[str, Any]:
    result = dict(row)
    raw = result["raw"]
    if isinstance(raw, str):
        raw = json.loads(raw)
    metadata = raw.get("_normalization", {})
    result["raw"] = raw
    result["status_note"] = metadata.get("status_note")
    result["reason_code"] = metadata.get("reason_code")
    result["reason_codes"] = metadata.get("reason_codes", {})
    result["vat_included"] = bool(metadata.get("vat_included"))
    result["has_tier"] = (
        result["price_cash"] is not None
        and result["price_term"] is not None
        and result["price_cash"] != result["price_term"]
    )
    if result["state"] == "REVIEW":
        result["change_kind"] = "REVIEW"
    elif result["state"] == "UNMATCHED":
        result["change_kind"] = "UNMATCHED"
    elif result["state"] == "SKIPPED":
        result["change_kind"] = "SKIPPED"
    elif result["old_currency"] is None:
        result["change_kind"] = "NEW_PRICE"
    elif any(
        result[current] != result[old]
        for current, old in (
            ("price_cash", "old_price_cash"),
            ("price_term", "old_price_term"),
            ("currency", "old_currency"),
            ("vat_rate", "old_vat_rate"),
        )
    ):
        result["change_kind"] = "CHANGED_PRICE"
    else:
        result["change_kind"] = "UNCHANGED"
    return result


@router.post("/imports/{import_id}/lines/{line_no}/override")
def override_line(import_id: int, line_no: int, payload: OverrideInput, request: Request,
                  db: Session = Depends(get_db)):
    _require(request, "supplier_prices.override_block")
    cid, actor = company_id(request), _actor(request)
    result = db.execute(text("""UPDATE supplier_price_import_lines SET override_by=:uid,
        override_at=:now,override_reason=:reason
        WHERE company_id=:cid AND import_id=:iid AND line_no=:line AND state='BLOCKED'
        AND override_at IS NULL"""),
        {"uid": actor.get("id"), "now": _now(), "reason": payload.reason,
         "cid": cid, "iid": import_id, "line": line_no})
    if result.rowcount != 1:
        existing = db.execute(text("""SELECT override_at FROM supplier_price_import_lines
            WHERE company_id=:cid AND import_id=:iid AND line_no=:line
            AND state='BLOCKED'"""),
            {"cid": cid, "iid": import_id, "line": line_no}).mappings().first()
        db.rollback()
        if existing and existing["override_at"] is not None:
            return {"ok": True}
        raise HTTPException(404, "Bloklu satır bulunamadı")
    log_activity(db, cid, actor.get("id"), "supplier_price.block_overridden",
                 "supplier_price_import", import_id, "Bloklu fiyat satırı kabul edildi",
                 {"line_no": line_no, "reason": payload.reason})
    db.commit()
    return {"ok": True}


def _price_rows(db: Session, cid: int, supplier_id: int) -> list[dict[str, Any]]:
    return [dict(row) for row in db.execute(text("""SELECT part_id,price_cash,price_term,currency,vat_rate
        FROM supplier_part_prices WHERE company_id=:cid AND supplier_id=:sid ORDER BY part_id"""),
        {"cid": cid, "sid": supplier_id}).mappings()]


def _same_snapshot(line: Any, current: Any) -> bool:
    if current is None:
        return all(
            line[field] is None
            for field in ("old_price_cash", "old_price_term", "old_currency", "old_vat_rate")
        )
    return all(
        line[line_field] == current[current_field]
        for line_field, current_field in (
            ("old_price_cash", "price_cash"),
            ("old_price_term", "price_term"),
            ("old_currency", "currency"),
            ("old_vat_rate", "vat_rate"),
        )
    )


@router.post("/imports/{import_id}/apply")
def apply_import(import_id: int, payload: ApplyInput, request: Request, db: Session = Depends(get_db)):
    _require(request, "supplier_prices.apply")
    cid, actor = company_id(request), _actor(request)
    is_postgresql = bool(db.bind and db.bind.dialect.name == "postgresql")
    lock = " FOR UPDATE" if is_postgresql else ""
    try:
        item = db.execute(text("""SELECT * FROM supplier_price_imports
            WHERE company_id=:cid AND id=:id""" + (" FOR UPDATE NOWAIT" if is_postgresql else "")),
            {"cid": cid, "id": import_id}).mappings().first()
    except OperationalError as exc:
        db.rollback()
        raise HTTPException(409, "Import eşzamanlı işleniyor") from exc
    if not item:
        raise HTTPException(404, "Import bulunamadı")
    if item["status"] != "DRY_RUN_READY" or payload.expected_revision != item["report_digest"] or payload.report_digest != item["report_digest"]:
        raise HTTPException(409, "Dry-run raporu güncel değil")
    remaining = db.execute(text("""SELECT COUNT(*) FROM supplier_price_import_lines
        WHERE company_id=:cid AND import_id=:id AND state='BLOCKED' AND override_at IS NULL"""),
        {"cid": cid, "id": import_id}).scalar_one()
    if remaining:
        raise HTTPException(409, "Kabul edilmemiş bloklu satırlar var")
    duplicate = db.execute(text("""SELECT id FROM supplier_price_imports
        WHERE company_id=:cid AND source_sha256=:sha AND status='APPLIED' AND id<>:id"""),
        {"cid": cid, "sha": item["source_sha256"], "id": import_id}).first()
    if duplicate:
        raise HTTPException(409, "Bu dosya daha önce uygulandı")
    lines = db.execute(text("""SELECT * FROM supplier_price_import_lines
        WHERE company_id=:cid AND import_id=:id AND part_id IS NOT NULL
        AND (state IN ('OK','WARNING') OR override_at IS NOT NULL) ORDER BY part_id"""),
        {"cid": cid, "id": import_id}).mappings().all()
    db.execute(text("SELECT id FROM companies WHERE id=:cid" + lock), {"cid": cid}).first()
    current_by_part: dict[int, Any] = {}
    for line in lines:
        current = db.execute(text("""SELECT * FROM supplier_part_prices
            WHERE company_id=:cid AND supplier_id=:sid AND part_id=:pid""" + lock),
            {"cid": cid, "sid": item["supplier_id"], "pid": line["part_id"]}).mappings().first()
        if not _same_snapshot(line, current):
            db.rollback()
            raise HTTPException(409, "Dry-run sonrasında tedarikçi fiyatı değişti")
        current_by_part[int(line["part_id"])] = current
    cas = db.execute(text("""UPDATE supplier_price_imports SET status='APPLYING'
        WHERE company_id=:cid AND id=:id AND status='DRY_RUN_READY'"""), {"cid": cid, "id": import_id})
    if cas.rowcount != 1:
        db.rollback()
        raise HTTPException(409, "Import eşzamanlı değiştirildi")
    now = _now()
    for line in lines:
        current = current_by_part[int(line["part_id"])]
        db.execute(text("""UPDATE supplier_price_import_lines SET old_price_cash=:cash,
            old_price_term=:term,old_currency=:currency,old_vat_rate=:vat
            WHERE company_id=:cid AND import_id=:iid AND id=:id"""),
            {"cash": current["price_cash"] if current else None, "term": current["price_term"] if current else None,
             "currency": current["currency"] if current else None, "vat": current["vat_rate"] if current else None,
             "cid": cid, "iid": import_id, "id": line["id"]})
        values = {"cash": line["price_cash"], "term": line["price_term"], "currency": line["currency"],
                  "vat": line["vat_rate"], "days": db.execute(text("SELECT term_days FROM supplier_import_profiles WHERE company_id=:cid AND id=:id"),
                  {"cid": cid, "id": item["profile_id"]}).scalar_one(), "source": import_id, "now": now,
                  "cid": cid, "sid": item["supplier_id"], "pid": line["part_id"]}
        if current:
            db.execute(text("""UPDATE supplier_part_prices SET price_cash=:cash,price_term=:term,
                currency=:currency,vat_rate=:vat,term_days=:days,source_import_id=:source,
                valid_from=:now,updated_at=:now WHERE company_id=:cid AND supplier_id=:sid AND part_id=:pid"""), values)
        else:
            db.execute(text("""INSERT INTO supplier_part_prices(company_id,supplier_id,part_id,
                price_cash,price_term,currency,vat_rate,term_days,source_import_id,valid_from,updated_at)
                VALUES(:cid,:sid,:pid,:cash,:term,:currency,:vat,:days,:source,:now,:now)"""), values)
    previous = db.execute(text("""SELECT COALESCE(MAX(revision),0) FROM supplier_price_imports
        WHERE company_id=:cid"""), {"cid": cid}).scalar_one()
    revision = int(previous) + 1
    digest = canonical_price_digest(_price_rows(db, cid, int(item["supplier_id"])))
    db.execute(text("""UPDATE supplier_price_imports SET status='APPLIED',revision=:revision,
        after_digest=:digest,applied_at=:now,applied_by=:uid WHERE company_id=:cid AND id=:id"""),
        {"revision": revision, "digest": digest, "now": now, "uid": actor.get("id"), "cid": cid, "id": import_id})
    log_activity(db, cid, actor.get("id"), "supplier_price.import_applied",
                 "supplier_price_import", import_id, "Tedarikçi fiyat importu uygulandı",
                 {"revision": revision, "rows": len(lines), "digest": digest})
    db.commit()
    return {"status": "APPLIED", "revision": revision, "after_digest": digest}


@router.post("/imports/{import_id}/revert")
def revert_import(import_id: int, request: Request, db: Session = Depends(get_db)):
    _require(request, "supplier_prices.apply")
    cid, actor = company_id(request), _actor(request)
    is_postgresql = bool(db.bind and db.bind.dialect.name == "postgresql")
    lock = " FOR UPDATE" if is_postgresql else ""
    try:
        item = db.execute(text("""SELECT * FROM supplier_price_imports
            WHERE company_id=:cid AND id=:id""" + (" FOR UPDATE NOWAIT" if is_postgresql else "")),
            {"cid": cid, "id": import_id}).mappings().first()
    except OperationalError as exc:
        db.rollback()
        raise HTTPException(409, "Import eşzamanlı işleniyor") from exc
    if not item:
        raise HTTPException(404, "Import bulunamadı")
    db.execute(text("SELECT id FROM companies WHERE id=:cid" + lock), {"cid": cid}).first()
    latest = db.execute(text("SELECT COALESCE(MAX(revision),0) FROM supplier_price_imports WHERE company_id=:cid"),
                        {"cid": cid}).scalar_one()
    current_digest = canonical_price_digest(_price_rows(db, cid, int(item["supplier_id"])))
    if item["status"] != "APPLIED" or item["revision"] != latest or current_digest != item["after_digest"]:
        raise HTTPException(409, "Aradan başka değişiklik geçmiş, otomatik geri alma güvenli değil")
    lines = db.execute(text("""SELECT * FROM supplier_price_import_lines
        WHERE company_id=:cid AND import_id=:id AND part_id IS NOT NULL
        AND (state IN ('OK','WARNING') OR override_at IS NOT NULL) ORDER BY part_id"""),
        {"cid": cid, "id": import_id}).mappings().all()
    now = _now()
    for line in lines:
        if line["old_price_cash"] is None and line["old_price_term"] is None and line["old_currency"] is None:
            db.execute(text("""DELETE FROM supplier_part_prices WHERE company_id=:cid
                AND supplier_id=:sid AND part_id=:pid"""),
                {"cid": cid, "sid": item["supplier_id"], "pid": line["part_id"]})
        else:
            db.execute(text("""UPDATE supplier_part_prices SET price_cash=:cash,price_term=:term,
                currency=:currency,vat_rate=:vat,updated_at=:now
                WHERE company_id=:cid AND supplier_id=:sid AND part_id=:pid"""),
                {"cash": line["old_price_cash"], "term": line["old_price_term"],
                 "currency": line["old_currency"], "vat": line["old_vat_rate"], "now": now,
                 "cid": cid, "sid": item["supplier_id"], "pid": line["part_id"]})
    revision = int(latest) + 1
    db.execute(text("""UPDATE supplier_price_imports SET status='REVERTED',revision=:revision,
        reverted_at=:now,reverted_by=:uid WHERE company_id=:cid AND id=:id"""),
        {"revision": revision, "now": now, "uid": actor.get("id"), "cid": cid, "id": import_id})
    log_activity(db, cid, actor.get("id"), "supplier_price.import_reverted",
                 "supplier_price_import", import_id, "Tedarikçi fiyat importu geri alındı",
                 {"revision": revision, "rows": len(lines)})
    db.commit()
    return {"status": "REVERTED", "revision": revision}


@router.post("/xrefs", status_code=201)
def create_xref(payload: XrefInput, request: Request, db: Session = Depends(get_db)):
    _require(request, "supplier_prices.import")
    cid, actor = company_id(request), _actor(request)
    owned = db.execute(text("""SELECT p.id FROM products p JOIN suppliers s ON s.id=:sid
        WHERE p.company_id=:cid AND p.id=:pid AND s.company_id=:cid"""),
        {"sid": payload.supplier_id, "cid": cid, "pid": payload.part_id}).first()
    if not owned:
        raise HTTPException(404, "Parça veya tedarikçi bulunamadı")
    xref_id = db.execute(text("""INSERT INTO supplier_part_xrefs(company_id,supplier_id,
        supplier_code,part_id,created_at,created_by)
        VALUES(:cid,:sid,:code,:pid,:now,:uid) RETURNING id"""),
        {"cid": cid, "sid": payload.supplier_id, "code": payload.supplier_code.strip(),
         "pid": payload.part_id, "now": _now(), "uid": actor.get("id")}).scalar_one()
    log_activity(db, cid, actor.get("id"), "supplier_price.xref_created",
                 "supplier_price_import", None, "Tedarikçi parça eşlemesi oluşturuldu",
                 {"supplier_id": payload.supplier_id, "supplier_code": payload.supplier_code,
                  "part_id": payload.part_id})
    db.commit()
    return {"id": xref_id}
