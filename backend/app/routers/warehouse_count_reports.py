from __future__ import annotations

import re
from datetime import datetime
from decimal import Decimal
from io import BytesIO

from fastapi import Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import text
from sqlalchemy.orm import Session

from ..auth import has_permission
from ..db import get_db
from ..money import quantity
from ..tenancy import company_id
from .outputs import router

_COUNT_NOTE = re.compile(
    r"^Fiziksel sayım \| Sistem: (?P<system>-?[0-9.]+) \| Sayılan: "
    r"(?P<counted>-?[0-9.]+)(?: \| Not: (?P<note>.*))?$"
)
_FORMULA_PREFIX = re.compile(r"^[\t\r\n ]*[=+\-@]")


def _date(value: str | None, label: str) -> str | None:
    if value is None or not value.strip():
        return None
    cleaned = value.strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(cleaned, fmt).date().isoformat()
        except ValueError:
            pass
    raise HTTPException(422, f"{label} YYYY-AA-GG veya GG.AA.YYYY biçiminde olmalıdır")


def _safe_cell(value):
    if isinstance(value, str) and _FORMULA_PREFIX.match(value):
        return "'" + value
    return value


def _parsed_quantities(note: str | None, variance: Decimal) -> tuple[Decimal, Decimal, str | None]:
    match = _COUNT_NOTE.match(note or "")
    if not match:
        counted = quantity(variance)
        return quantity(0), counted, note
    return quantity(match.group("system")), quantity(match.group("counted")), match.group("note") or None


def _require_stock_permission(request: Request) -> None:
    user = getattr(request.state, "user", None)
    role = str(user.get("role", "")) if isinstance(user, dict) else ""
    if not has_permission(role, "stock"):
        raise HTTPException(403, "Bu rapor için stok yetkisi gerekiyor")


@router.get("/exports/warehouse-count-variance.xlsx")
def variance_report_xlsx(
    request: Request,
    warehouse_id: int | None = Query(default=None, gt=0),
    date_from: str | None = None,
    date_to: str | None = None,
    changed_only: bool = False,
    min_absolute_variance: Decimal = Query(default=Decimal("0"), ge=0),
    db: Session = Depends(get_db),
):
    """Export tenant-scoped physical count documents and their variance lines."""
    _require_stock_permission(request)
    cid = company_id(request)
    start = _date(date_from, "Başlangıç tarihi")
    end = _date(date_to, "Bitiş tarihi")
    threshold = quantity(min_absolute_variance)
    if start and end and start > end:
        raise HTTPException(422, "Başlangıç tarihi bitiş tarihinden sonra olamaz")

    if warehouse_id is not None:
        exists = db.execute(
            text("SELECT 1 FROM warehouses WHERE id=:wid AND company_id=:cid"),
            {"wid": warehouse_id, "cid": cid},
        ).scalar()
        if not exists:
            raise HTTPException(404, "Depo bulunamadı")

    clauses = [
        "sm.company_id=:cid",
        "sm.reference_type='inventory_count'",
        "sm.reference_id IS NOT NULL",
    ]
    params: dict[str, object] = {"cid": cid}
    if warehouse_id is not None:
        clauses.append("sm.warehouse_id=:wid")
        params["wid"] = warehouse_id
    if start:
        clauses.append("sm.movement_date>=:date_from")
        params["date_from"] = start
    if end:
        clauses.append("sm.movement_date<=:date_to")
        params["date_to"] = end
    if changed_only:
        clauses.append("sm.quantity<>0")
    if threshold > 0:
        clauses.append(
            """(SELECT COALESCE(SUM(ABS(sm_total.quantity)),0)
            FROM stock_movements sm_total
            WHERE sm_total.company_id=sm.company_id
              AND sm_total.reference_type='inventory_count'
              AND sm_total.reference_id=sm.reference_id)>=:min_absolute_variance"""
        )
        params["min_absolute_variance"] = threshold

    rows = db.execute(
        text(
            f"""SELECT sm.reference_id count_id,sm.movement_date count_date,
            sm.warehouse_id,w.name warehouse_name,sm.product_id,p.name product_name,
            p.product_code,p.unit,sm.quantity variance,sm.note
            FROM stock_movements sm
            JOIN warehouses w ON w.id=sm.warehouse_id AND w.company_id=sm.company_id
            JOIN products p ON p.id=sm.product_id AND p.company_id=sm.company_id
            WHERE {' AND '.join(clauses)}
            ORDER BY sm.movement_date DESC,sm.reference_id DESC,p.name,sm.id"""
        ),
        params,
    ).mappings().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Sayım Farkları"
    headers = [
        "Sayım No", "Tarih", "Depo", "Ürün", "Ürün Kodu", "Birim",
        "Sistem Miktarı", "Sayılan Miktar", "Fark", "Yön", "Açıklama",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="17324D")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    total_variance = Decimal("0")
    changed_lines = 0
    for row in rows:
        variance = quantity(row["variance"])
        system_quantity, counted_quantity, user_note = _parsed_quantities(row["note"], variance)
        total_variance += abs(variance)
        changed_lines += int(variance != 0)
        direction = "Artış" if variance > 0 else "Azalış" if variance < 0 else "Fark Yok"
        ws.append([
            row["count_id"], row["count_date"], _safe_cell(row["warehouse_name"]),
            _safe_cell(row["product_name"]), _safe_cell(row["product_code"]), _safe_cell(row["unit"]),
            system_quantity, counted_quantity, variance, direction,
            _safe_cell(user_note),
        ])

    ws.auto_filter.ref = f"A1:K{max(1, ws.max_row)}"
    ws.freeze_panes = "A2"
    widths = [12, 14, 24, 36, 18, 12, 18, 18, 14, 12, 40]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + index)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top")

    summary = wb.create_sheet("Özet")
    summary.append(["FİZİKSEL SAYIM FARK RAPORU"])
    summary["A1"].font = Font(size=16, bold=True)
    summary.append([])
    summary.append(["Toplam satır", len(rows)])
    summary.append(["Fark bulunan satır", changed_lines])
    summary.append(["Toplam mutlak fark", total_variance])
    summary.append(["Depo filtresi", warehouse_id or "Tüm depolar"])
    summary.append(["Başlangıç", start or "-"])
    summary.append(["Bitiş", end or "-"])
    summary.append(["Yalnız fark bulunanlar", "Evet" if changed_only else "Hayır"])
    summary.append(["Asgari belge mutlak farkı", threshold])
    summary.column_dimensions["A"].width = 30
    summary.column_dimensions["B"].width = 24

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="stok-sayim-farklari.xlsx"'},
    )
