from __future__ import annotations

import asyncio
from decimal import Decimal
from io import BytesIO

import pytest
from fastapi import HTTPException
from openpyxl import load_workbook
from starlette.requests import Request

from app.routers.warehouse_count_reports import variance_report_xlsx


class Result:
    def __init__(self, *, scalar_value=None, rows=None):
        self.scalar_value = scalar_value
        self.rows = rows or []

    def scalar(self):
        return self.scalar_value

    def mappings(self):
        return self

    def all(self):
        return self.rows


class Session:
    def __init__(self, warehouse_exists=True):
        self.warehouse_exists = warehouse_exists
        self.calls = []

    def execute(self, statement, params):
        sql = str(statement)
        self.calls.append((sql, dict(params)))
        if "SELECT 1 FROM warehouses" in sql:
            return Result(scalar_value=1 if self.warehouse_exists else None)
        return Result(rows=[{
            "count_id": 41,
            "count_date": "2026-07-20",
            "warehouse_id": 3,
            "warehouse_name": "  =Merkez",
            "product_id": 10,
            "product_name": "+Pompa",
            "product_code": "@HP-10",
            "unit": "@Adet",
            "variance": Decimal("-3"),
            "note": "Fiziksel sayım | Sistem: 10 | Sayılan: 7 | Not: -Raf A",
        }])


def request(company_id=7, role="admin"):
    req = Request({"type": "http", "method": "GET", "path": "/", "headers": []})
    req.state.company_id = company_id
    req.state.user = {"role": role}
    return req


async def body(response):
    chunks = []
    async for chunk in response.body_iterator:
        chunks.append(chunk.encode() if isinstance(chunk, str) else chunk)
    return b"".join(chunks)


def test_variance_report_workbook_is_tenant_scoped_and_filter_bound() -> None:
    db = Session()
    response = variance_report_xlsx(
        request(), warehouse_id=3, date_from="2026-07-01",
        date_to="20.07.2026", changed_only=True,
        min_absolute_variance=Decimal("3"), db=db,
    )
    workbook = load_workbook(BytesIO(asyncio.run(body(response))), data_only=False)

    assert response.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert workbook.sheetnames == ["Sayım Farkları", "Özet"]
    sheet = workbook["Sayım Farkları"]
    assert sheet.freeze_panes == "A2"
    assert sheet["C2"].value == "'  =Merkez"
    assert sheet["D2"].value == "'+Pompa"
    assert sheet["E2"].value == "'@HP-10"
    assert sheet["F2"].value == "'@Adet"
    assert sheet["G2"].value == 10
    assert sheet["H2"].value == 7
    assert sheet["I2"].value == -3
    assert sheet["J2"].value == "Azalış"
    assert sheet["K2"].value == "'-Raf A"

    summary = workbook["Özet"]
    assert summary["A10"].value == "Asgari belge mutlak farkı"
    assert summary["B3"].value == 1
    assert summary["B4"].value == 1
    assert summary["B5"].value == 3
    assert summary["B10"].value == 3

    warehouse_sql, warehouse_params = db.calls[0]
    report_sql, report_params = db.calls[1]
    assert "company_id=:cid" in warehouse_sql
    assert warehouse_params == {"wid": 3, "cid": 7}
    assert "sm.company_id=:cid" in report_sql
    assert "sm.warehouse_id=:wid" in report_sql
    assert "sm.movement_date>=:date_from" in report_sql
    assert "sm.movement_date<=:date_to" in report_sql
    assert "sm.quantity<>0" in report_sql
    assert "SUM(ABS(sm_total.quantity))" in report_sql
    assert "sm_total.company_id=sm.company_id" in report_sql
    assert "sm_total.reference_id=sm.reference_id" in report_sql
    assert ">=:min_absolute_variance" in report_sql
    assert report_params == {
        "cid": 7, "wid": 3,
        "date_from": "2026-07-01", "date_to": "2026-07-20",
        "min_absolute_variance": Decimal("3.0000"),
    }


def test_variance_report_requires_stock_permission() -> None:
    with pytest.raises(HTTPException) as forbidden:
        variance_report_xlsx(
            request(role="satis"), warehouse_id=None, date_from=None,
            date_to=None, changed_only=False,
            min_absolute_variance=Decimal("0"), db=Session(),
        )
    assert forbidden.value.status_code == 403


def test_variance_report_rejects_invalid_range_and_foreign_warehouse() -> None:
    with pytest.raises(HTTPException) as invalid:
        variance_report_xlsx(
            request(), warehouse_id=None, date_from="2026-07-21",
            date_to="2026-07-20", changed_only=False,
            min_absolute_variance=Decimal("0"), db=Session(),
        )
    assert invalid.value.status_code == 422

    with pytest.raises(HTTPException) as foreign:
        variance_report_xlsx(
            request(), warehouse_id=999, date_from=None,
            date_to=None, changed_only=False,
            min_absolute_variance=Decimal("0"), db=Session(False),
        )
    assert foreign.value.status_code == 404
